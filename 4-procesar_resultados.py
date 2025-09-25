# procesar_resultados.py (Versión Actualizada con Tiempos de Competencia)

import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

# --- CONFIGURACIÓN ---
ARCHIVO_ENTRADA_RESULTADOS = 'resultados_con_tiempos.xlsx'
ARCHIVO_SALIDA_PREMIACION = 'reporte_premiacion_final_CORREGIDO.xlsx'

# Nuevo sistema de puntos: 1º=9, 2º=7, 3º=6, 4º=5, luego -1 por posición
def calcular_puntos(posicion):
    if posicion == 1:
        return 9
    elif posicion == 2:
        return 7
    elif posicion == 3:
        return 6
    elif posicion == 4:
        return 5
    else:
        # A partir del 5º lugar, se resta 1 punto por posición (4, 3, 2, 1)
        puntos = max(1, 5 - (posicion - 4))
        return puntos

def parse_time(time_val):
    """Convierte tiempo a segundos para ordenamiento"""
    if pd.isna(time_val) or time_val == "" or time_val is None:
        return float('inf')

    if hasattr(time_val, 'minute'):
        return time_val.minute * 60 + time_val.second + time_val.microsecond / 1_000_000

    time_str = str(time_val).replace(',', '.')
    try:
        parts = time_str.split(':')
        if len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
        return float(time_str)
    except (ValueError, IndexError):
        return float('inf')

def leer_tiempos_competencia_desde_sembrado():
    """Lee los tiempos de competencia desde los archivos de sembrado manual"""
    todos_resultados = []

    # Buscar todos los archivos de sembrado manual en session_state
    import streamlit as st

    if 'manual_seeding_' not in str(st.session_state.keys()):
        return []

    # Buscar todas las claves de sembrado manual
    for key in st.session_state.keys():
        if key.startswith('manual_seeding_'):
            seeding_data = st.session_state[key]

            # Extraer evento y género del key
            parts = key.replace('manual_seeding_', '').rsplit('_', 1)
            if len(parts) == 2:
                evento = parts[0]
                genero = 'Hombres' if parts[1] == 'M' else 'Mujeres'

                # Procesar cada serie
                for serie in seeding_data.get('series', []):
                    for carril_idx, swimmer in enumerate(serie['carriles']):
                        if swimmer and swimmer.get('tiempo_competencia'):
                            # Solo incluir nadadores con tiempo de competencia
                            resultado = {
                                'Prueba': f"{evento} - {genero}",
                                'Evento': evento,
                                'Genero': genero,
                                'Serie': serie['serie'],
                                'Carril': carril_idx + 1,
                                'Nombre': swimmer['nombre'],
                                'Equipo': swimmer['equipo'],
                                'Edad': swimmer['edad'],
                                'Categoria': swimmer['categoria'],
                                'Tiempo_Sembrado': swimmer['tiempo'],
                                'Tiempo_Competencia': swimmer['tiempo_competencia'],
                                'Tiempo_Segundos': parse_time(swimmer['tiempo_competencia'])
                            }
                            todos_resultados.append(resultado)

    return todos_resultados

def procesar_resultados_por_categoria_y_genero(resultados):
    """Procesa resultados agrupados por categoría y género con nuevo sistema de puntos"""
    if not resultados:
        return []

    df = pd.DataFrame(resultados)
    resultados_procesados = []

    # Agrupar por evento, género y categoría
    for (evento, genero, categoria), grupo in df.groupby(['Evento', 'Genero', 'Categoria']):
        # Ordenar por tiempo (menor a mayor)
        grupo_ordenado = grupo.sort_values('Tiempo_Segundos')

        for posicion, (_, nadador) in enumerate(grupo_ordenado.iterrows(), 1):
            # Solo asignar puntos si el tiempo es válido (no inf)
            if nadador['Tiempo_Segundos'] != float('inf'):
                puntos = calcular_puntos(posicion)

                resultado = {
                    'Evento': evento,
                    'Genero': genero,
                    'Categoria': categoria,
                    'Posicion': posicion,
                    'Nombre': nadador['Nombre'],
                    'Equipo': nadador['Equipo'],
                    'Edad': nadador['Edad'],
                    'Tiempo_Competencia': nadador['Tiempo_Competencia'],
                    'Puntos': puntos
                }
                resultados_procesados.append(resultado)

    return resultados_procesados

def generar_resumen_equipos(resultados_procesados):
    """Genera resumen de puntos por equipo"""
    if not resultados_procesados:
        return []

    df = pd.DataFrame(resultados_procesados)
    resumen_equipos = df.groupby('Equipo').agg({
        'Puntos': 'sum',
        'Nombre': 'count'  # Contar participaciones
    }).reset_index()

    resumen_equipos.columns = ['Equipo', 'Puntos_Total', 'Participaciones']
    resumen_equipos = resumen_equipos.sort_values('Puntos_Total', ascending=False)

    return resumen_equipos.to_dict('records')

def generar_reporte_resultados_completo():
    """Función principal que genera el reporte completo de resultados"""
    try:
        # 1. Leer tiempos de competencia desde sembrado manual
        resultados_brutos = leer_tiempos_competencia_desde_sembrado()

        if not resultados_brutos:
            return False, "No se encontraron tiempos de competencia en el sembrado manual"

        # 2. Procesar resultados por categoría y género
        resultados_procesados = procesar_resultados_por_categoria_y_genero(resultados_brutos)

        if not resultados_procesados:
            return False, "No se pudieron procesar los resultados"

        # 3. Generar resumen por equipos
        resumen_equipos = generar_resumen_equipos(resultados_procesados)

        # 4. Crear archivo Excel con múltiples hojas
        wb = Workbook()

        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # HOJA 1: Resultados por Categoría y Género
        ws_resultados = wb.create_sheet("Resultados por Categoría")

        # Encabezados para resultados
        headers_resultados = ['Evento', 'Género', 'Categoría', 'Pos.', 'Nombre', 'Equipo', 'Edad', 'Tiempo', 'Puntos']
        for col, header in enumerate(headers_resultados, 1):
            cell = ws_resultados.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")

        # Llenar datos de resultados
        row = 2
        for resultado in resultados_procesados:
            ws_resultados.cell(row=row, column=1, value=resultado['Evento'])
            ws_resultados.cell(row=row, column=2, value=resultado['Genero'])
            ws_resultados.cell(row=row, column=3, value=resultado['Categoria'])
            ws_resultados.cell(row=row, column=4, value=resultado['Posicion'])
            ws_resultados.cell(row=row, column=5, value=resultado['Nombre'])
            ws_resultados.cell(row=row, column=6, value=resultado['Equipo'])
            ws_resultados.cell(row=row, column=7, value=resultado['Edad'])
            ws_resultados.cell(row=row, column=8, value=resultado['Tiempo_Competencia'])
            ws_resultados.cell(row=row, column=9, value=resultado['Puntos'])

            # Colorear las medallas
            if resultado['Posicion'] == 1:
                for col in range(1, 10):
                    ws_resultados.cell(row=row, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif resultado['Posicion'] == 2:
                for col in range(1, 10):
                    ws_resultados.cell(row=row, column=col).fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            elif resultado['Posicion'] == 3:
                for col in range(1, 10):
                    ws_resultados.cell(row=row, column=col).fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

            row += 1

        # Ajustar anchos de columna
        column_widths = [15, 10, 12, 5, 25, 20, 5, 12, 8]
        for i, width in enumerate(column_widths, 1):
            ws_resultados.column_dimensions[chr(64+i)].width = width

        # HOJA 2: Clasificación por Equipos
        ws_equipos = wb.create_sheet("Clasificación por Equipos")

        # Encabezados para equipos
        headers_equipos = ['Posición', 'Equipo', 'Puntos Totales', 'Participaciones']
        for col, header in enumerate(headers_equipos, 1):
            cell = ws_equipos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")

        # Llenar datos de equipos
        for pos, equipo in enumerate(resumen_equipos, 1):
            ws_equipos.cell(row=pos+1, column=1, value=pos)
            ws_equipos.cell(row=pos+1, column=2, value=equipo['Equipo'])
            ws_equipos.cell(row=pos+1, column=3, value=equipo['Puntos_Total'])
            ws_equipos.cell(row=pos+1, column=4, value=equipo['Participaciones'])

            # Colorear los 3 primeros equipos
            if pos == 1:
                for col in range(1, 5):
                    ws_equipos.cell(row=pos+1, column=col).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif pos == 2:
                for col in range(1, 5):
                    ws_equipos.cell(row=pos+1, column=col).fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            elif pos == 3:
                for col in range(1, 5):
                    ws_equipos.cell(row=pos+1, column=col).fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

        # Ajustar anchos de columna para equipos
        ws_equipos.column_dimensions['A'].width = 10
        ws_equipos.column_dimensions['B'].width = 25
        ws_equipos.column_dimensions['C'].width = 15
        ws_equipos.column_dimensions['D'].width = 15

        # Guardar archivo
        wb.save(ARCHIVO_SALIDA_PREMIACION)

        return True, f"Reporte generado exitosamente: {ARCHIVO_SALIDA_PREMIACION} ({len(resultados_procesados)} resultados, {len(resumen_equipos)} equipos)"

    except Exception as e:
        return False, f"Error al generar reporte: {str(e)}"

def main():
    """Función principal para ejecutar el procesamiento"""
    success, message = generar_reporte_resultados_completo()
    if success:
        print(f"✅ {message}")
    else:
        print(f"❌ {message}")
    return success

if __name__ == "__main__":
    main()

def procesar_resultados_excel_corregido(filepath):
    """Lee el archivo de resultados que ahora incluye la columna 'Categoría'."""
    try:
        df_raw = pd.read_excel(filepath, header=None)
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{filepath}'.")
        return None
    
    all_results = []
    current_prueba = ""

    for index, row in df_raw.iterrows():
        # Detectar el nombre de la prueba
        if " - " in str(row[0]) and pd.isna(row[1]):
            current_prueba = row[0]
            continue
        
        # Detectar la cabecera de la tabla de resultados
        if str(row[0]).strip() == "Carril":
            # Empezar a leer los nadadores de esta tabla
            for i in range(index + 1, len(df_raw)):
                nadador_row = df_raw.iloc[i]
                # Parar si la fila está vacía o es una nueva cabecera de serie/categoría
                if pd.isna(nadador_row[0]) or "Serie" in str(nadador_row[0]) or "Categoría" in str(nadador_row[0]):
                    break
                
                # <-- ¡CORRECCIÓN CLAVE! Leemos los datos por posición de columna
                # Col A(0): Carril, B(1): Nombre, C(2): Equipo, D(3): Edad, E(4): Categoría, G(6): Tiempo Final
                if pd.notna(nadador_row[1]): # Si hay un nombre
                    all_results.append({
                        "Prueba": current_prueba,
                        "Nombre": nadador_row[1],
                        "Equipo": nadador_row[2],
                        "Edad": nadador_row[3],
                        "Categoria": nadador_row[4], # <-- Leemos la categoría de su propia columna
                        "Tiempo Final": nadador_row[6] if len(nadador_row) > 6 else None
                    })
    return pd.DataFrame(all_results)

# ... [La función main y las de ayuda deben ser copiadas de la respuesta anterior, pero usando esta nueva función de lectura] ...

if __name__ == "__main__":
    # Copia aquí las funciones auxiliares (parse_time, format_time_value, apply_styles_and_width)
    # y la función main() completa de la respuesta anterior, pero asegúrate de que llame a:
    # df_results = procesar_resultados_excel_corregido(ARCHIVO_ENTRADA_RESULTADOS)
    def main():
        print("Iniciando el procesamiento de resultados (versión final corregida)...")
        
        # Llama a la nueva función de lectura
        df_results = procesar_resultados_excel_corregido(ARCHIVO_ENTRADA_RESULTADOS)

        if df_results is None or df_results.empty:
            print("No se encontraron datos para procesar. Finalizando.")
            return

        # El resto del proceso es idéntico al de la respuesta anterior...
        df_results['tiempo_final_segundos'] = df_results['Tiempo Final'].apply(parse_time)
        df_results['Sexo'] = df_results['Prueba'].apply(lambda x: 'F' if 'Mujeres' in x else 'M')
        df_results['Lugar'] = df_results.groupby(['Prueba', 'Categoria'])['tiempo_final_segundos'].rank(method='min').astype(int)
        df_results['Puntos'] = df_results['Lugar'].map(PUNTOS).fillna(0).astype(int)
        
        # ... (copia el resto de la función main de la respuesta anterior aquí para generar los 3 reportes) ...
        print(f"¡Éxito! Reporte final '{ARCHIVO_SALIDA_PREMIACION}' generado correctamente.")
    
    # Para que este script sea autocontenido, aquí está el main completo y las funciones auxiliares
    def format_time_value(time_val):
        if pd.isna(time_val): return 'N/A'
        if hasattr(time_val, 'strftime'): return time_val.strftime('%M:%S.%f')[:-4]
        return str(time_val)

    def apply_styles_and_width(ws):
        from openpyxl.styles import Font, Border, Side
        font_header = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = border
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    def parse_time(time_val):
        if pd.isna(time_val) or time_val == '': return float('inf')
        if hasattr(time_val, 'minute'): return time_val.minute * 60 + time_val.second + time_val.microsecond / 1_000_000
        time_str = str(time_val).replace(',', '.')
        try:
            parts = time_str.split(':')
            if len(parts) == 2: return int(parts[0]) * 60 + float(parts[1])
            return float(time_str)
        except (ValueError, IndexError): return float('inf')


    # --- FUNCIÓN MAIN COMPLETA ---
    def main_full():
        print("Iniciando el procesamiento de resultados (versión final corregida)...")
        
        df_results = procesar_resultados_excel_corregido(ARCHIVO_ENTRADA_RESULTADOS)
        if df_results is None or df_results.empty:
            print("No se encontraron datos para procesar. Finalizando.")
            return

        df_results['tiempo_final_segundos'] = df_results['Tiempo Final'].apply(parse_time)
        df_results['Sexo'] = df_results['Prueba'].apply(lambda x: 'F' if 'Mujeres' in x else 'M')
        df_results['Lugar'] = df_results.groupby(['Prueba', 'Categoria'])['tiempo_final_segundos'].rank(method='min').astype(int)
        df_results['Puntos'] = df_results['Lugar'].map(PUNTOS).fillna(0).astype(int)
        
        # Reportes individuales ordenados por tiempo (mejor tiempo = lugar 1)
        df_individual = df_results.groupby(['Categoria', 'Sexo', 'Nombre', 'Equipo']).agg({
            'Puntos': 'sum',
            'tiempo_final_segundos': 'min'  # Mejor tiempo del nadador
        }).rename(columns={'Puntos': 'Puntos_Totales'}).reset_index()
        
        # Ranking por puntos dentro de cada categoría y sexo
        df_individual['Lugar_Puntos'] = df_individual.groupby(['Categoria', 'Sexo'])['Puntos_Totales'].rank(method='min', ascending=False).astype(int)
        # Ranking por tiempo dentro de cada categoría y sexo
        df_individual['Lugar_Tiempo'] = df_individual.groupby(['Categoria', 'Sexo'])['tiempo_final_segundos'].rank(method='min', ascending=True).astype(int)
        
        reporte_individual = df_individual.sort_values(by=['Categoria', 'Sexo', 'Lugar_Tiempo'])

        # Reportes por equipo ordenados por puntos y tiempo promedio
        df_clubes = df_results.groupby('Equipo').agg({
            'Puntos': 'sum',
            'tiempo_final_segundos': 'mean'  # Tiempo promedio del equipo
        }).rename(columns={'Puntos': 'Puntos_Totales', 'tiempo_final_segundos': 'Tiempo_Promedio'}).reset_index()
        
        df_clubes['Lugar_Puntos'] = df_clubes['Puntos_Totales'].rank(method='min', ascending=False).astype(int)
        df_clubes['Lugar_Tiempo'] = df_clubes['Tiempo_Promedio'].rank(method='min', ascending=True).astype(int)
        reporte_clubes = df_clubes.sort_values(by='Lugar_Puntos')[['Lugar_Puntos', 'Equipo', 'Puntos_Totales', 'Tiempo_Promedio', 'Lugar_Tiempo']]

        wb = Workbook()
        
        # Pestaña 1: Ranking por Categoría (ordenado por tiempo)
        ws1 = wb.active
        ws1.title = "Ranking por Categoría"
        fila_actual = 1
        font_titulo_prueba = Font(bold=True, size=16)
        font_titulo_cat = Font(bold=True, size=14)
        font_header_tabla = Font(bold=True)
        
        ws1.cell(row=fila_actual, column=1, value="RANKING GENERAL POR CATEGORÍA (Ordenado por Tiempo)").font = font_titulo_prueba
        fila_actual += 3
        
        for categoria in sorted(df_results['Categoria'].unique()):
            ws1.cell(row=fila_actual, column=1, value=f"CATEGORÍA: {categoria}").font = font_titulo_cat
            fila_actual += 2
            
            df_categoria = df_results[df_results['Categoria'] == categoria].copy()
            # Ordenar por tiempo (mejor tiempo primero)
            df_categoria = df_categoria.sort_values(by='tiempo_final_segundos')
            df_categoria['Posicion_Categoria'] = range(1, len(df_categoria) + 1)
            
            df_categoria_display = df_categoria[['Posicion_Categoria', 'Nombre', 'Equipo', 'Sexo', 'Prueba', 'Tiempo Final', 'Puntos']]
            df_categoria_display['Tiempo Final'] = df_categoria_display['Tiempo Final'].apply(format_time_value)
            df_categoria_display['Sexo'] = df_categoria_display['Sexo'].map({'M': 'Masculino', 'F': 'Femenino'})
            
            headers = ['Lugar', 'Nombre', 'Equipo', 'Sexo', 'Prueba', 'Tiempo', 'Puntos']
            for col, header_text in enumerate(headers, 1):
                ws1.cell(row=fila_actual, column=col, value=header_text).font = font_header_tabla
            fila_actual += 1

            for _, r in df_categoria_display.iterrows():
                ws1.append(list(r))
            fila_actual += len(df_categoria_display) + 2
        apply_styles_and_width(ws1)

        # Pestaña 2: Ranking por Sexo (ordenado por tiempo)
        ws2 = wb.create_sheet("Ranking por Sexo")
        fila_actual = 1
        ws2.cell(row=fila_actual, column=1, value="RANKING GENERAL POR SEXO (Ordenado por Tiempo)").font = font_titulo_prueba
        fila_actual += 3
        
        for sexo in ['M', 'F']:
            sexo_texto = "MASCULINO" if sexo == 'M' else "FEMENINO"
            ws2.cell(row=fila_actual, column=1, value=f"CATEGORÍA: {sexo_texto}").font = font_titulo_cat
            fila_actual += 2
            
            df_sexo = df_results[df_results['Sexo'] == sexo].copy()
            # Ordenar por tiempo (mejor tiempo primero)
            df_sexo = df_sexo.sort_values(by='tiempo_final_segundos')
            df_sexo['Posicion_Sexo'] = range(1, len(df_sexo) + 1)
            
            df_sexo_display = df_sexo[['Posicion_Sexo', 'Nombre', 'Equipo', 'Categoria', 'Prueba', 'Tiempo Final', 'Puntos']]
            df_sexo_display['Tiempo Final'] = df_sexo_display['Tiempo Final'].apply(format_time_value)
            
            headers = ['Lugar', 'Nombre', 'Equipo', 'Categoría', 'Prueba', 'Tiempo', 'Puntos']
            for col, header_text in enumerate(headers, 1):
                ws2.cell(row=fila_actual, column=col, value=header_text).font = font_header_tabla
            fila_actual += 1
            
            for _, r in df_sexo_display.iterrows():
                ws2.append(list(r))
            fila_actual += len(df_sexo_display) + 2
        apply_styles_and_width(ws2)

        # Pestaña 3: Ranking por Equipo (ordenado por puntos y tiempo promedio)
        ws3 = wb.create_sheet("Ranking por Equipo")
        ws3.cell(row=1, column=1, value="RANKING GENERAL POR EQUIPOS (Ordenado por Puntos)").font = font_titulo_prueba
        fila_actual = 3
        
        # Tabla general de equipos
        df_equipos_display = reporte_clubes.copy()
        df_equipos_display['Tiempo_Promedio'] = df_equipos_display['Tiempo_Promedio'].apply(lambda x: f"{x:.2f}s" if pd.notna(x) else 'N/A')
        
        headers = ['Lugar', 'Equipo', 'Puntos Totales', 'Tiempo Promedio', 'Lugar por Tiempo']
        for col, header_text in enumerate(headers, 1):
            ws3.cell(row=fila_actual, column=col, value=header_text).font = font_header_tabla
        fila_actual += 1
        
        for _, r in df_equipos_display.iterrows():
            ws3.append(list(r))
        fila_actual += len(df_equipos_display) + 3
        
        # Detalle por equipo
        ws3.cell(row=fila_actual, column=1, value="DETALLE POR EQUIPOS (Ordenado por Tiempo)").font = font_titulo_cat
        fila_actual += 2
        
        for equipo in sorted(df_results['Equipo'].unique()):
            ws3.cell(row=fila_actual, column=1, value=f"EQUIPO: {equipo}").font = font_titulo_cat
            fila_actual += 1
            
            df_equipo = df_results[df_results['Equipo'] == equipo].copy()
            df_equipo = df_equipo.sort_values(by='tiempo_final_segundos')
            df_equipo['Posicion_Equipo'] = range(1, len(df_equipo) + 1)
            
            df_equipo_display = df_equipo[['Posicion_Equipo', 'Nombre', 'Categoria', 'Sexo', 'Prueba', 'Tiempo Final', 'Puntos']]
            df_equipo_display['Tiempo Final'] = df_equipo_display['Tiempo Final'].apply(format_time_value)
            df_equipo_display['Sexo'] = df_equipo_display['Sexo'].map({'M': 'M', 'F': 'F'})
            
            headers = ['Lugar', 'Nombre', 'Categoría', 'Sexo', 'Prueba', 'Tiempo', 'Puntos']
            for col, header_text in enumerate(headers, 1):
                ws3.cell(row=fila_actual, column=col, value=header_text).font = font_header_tabla
            fila_actual += 1
            
            for _, r in df_equipo_display.iterrows():
                ws3.append(list(r))
            fila_actual += len(df_equipo_display) + 2
            
        apply_styles_and_width(ws3)

        try:
            wb.save(ARCHIVO_SALIDA_PREMIACION)
            print("-" * 50)
            print("¡Proceso completado con éxito!")
            print(f"El reporte de premiación mejorado ha sido generado en:")
            print(f"'{ARCHIVO_SALIDA_PREMIACION}'")
            print("-" * 50)
        except Exception as e:
            print(f"Error al guardar el archivo de salida: {e}")

    def get_resultados_data():
        """Función para obtener datos procesados para Streamlit"""
        try:
            df_results = procesar_resultados_excel_corregido(ARCHIVO_ENTRADA_RESULTADOS)
            if df_results is None or df_results.empty:
                return None, "No se encontraron datos para procesar"

            df_results['tiempo_final_segundos'] = df_results['Tiempo Final'].apply(parse_time)
            df_results['Sexo'] = df_results['Prueba'].apply(lambda x: 'F' if 'Mujeres' in x else 'M')
            df_results['Lugar'] = df_results.groupby(['Prueba', 'Categoria'])['tiempo_final_segundos'].rank(method='min').astype(int)
            df_results['Puntos'] = df_results['Lugar'].map(PUNTOS).fillna(0).astype(int)
            
            # Formatear tiempo para display
            df_results['Tiempo_Formateado'] = df_results['Tiempo Final'].apply(format_time_value)
            df_results['Sexo_Display'] = df_results['Sexo'].map({'M': 'Masculino', 'F': 'Femenino'})
            
            # Crear datos por categoría
            datos_categoria = {}
            for categoria in sorted(df_results['Categoria'].unique()):
                df_cat = df_results[df_results['Categoria'] == categoria].sort_values(by='tiempo_final_segundos')
                df_cat['Posicion'] = range(1, len(df_cat) + 1)
                datos_categoria[categoria] = df_cat
            
            # Crear datos por sexo
            datos_sexo = {}
            for sexo in ['M', 'F']:
                sexo_name = 'Masculino' if sexo == 'M' else 'Femenino'
                df_sex = df_results[df_results['Sexo'] == sexo].sort_values(by='tiempo_final_segundos')
                df_sex['Posicion'] = range(1, len(df_sex) + 1)
                datos_sexo[sexo_name] = df_sex
            
            # Crear datos por equipo
            datos_equipo = {}
            # Resumen de equipos
            df_equipos_resumen = df_results.groupby('Equipo').agg({
                'Puntos': 'sum',
                'tiempo_final_segundos': 'mean'
            }).rename(columns={'Puntos': 'Puntos_Totales', 'tiempo_final_segundos': 'Tiempo_Promedio'}).reset_index()
            df_equipos_resumen['Lugar'] = df_equipos_resumen['Puntos_Totales'].rank(method='min', ascending=False).astype(int)
            df_equipos_resumen = df_equipos_resumen.sort_values(by='Lugar')
            datos_equipo['_resumen'] = df_equipos_resumen
            
            # Detalle por cada equipo
            for equipo in sorted(df_results['Equipo'].unique()):
                df_eq = df_results[df_results['Equipo'] == equipo].sort_values(by='tiempo_final_segundos')
                df_eq['Posicion'] = range(1, len(df_eq) + 1)
                datos_equipo[equipo] = df_eq
            
            return {
                'por_categoria': datos_categoria,
                'por_sexo': datos_sexo, 
                'por_equipo': datos_equipo,
                'df_completo': df_results
            }, "Datos procesados correctamente"
            
        except Exception as e:
            return None, f"Error al procesar datos: {e}"

    main_full()