import streamlit as st
import pandas as pd
import io
import os
from pathlib import Path
from datetime import datetime
#import subprocess  # No longer needed
import sys

# Importar las funciones de los scripts existentes
#import importlib.util # No longer needed

# Importar los scripts directly  
import importlib.util
from planilla_utils import inscrito_en_prueba

# Importar el módulo de inscripción con el nuevo nombre
spec = importlib.util.spec_from_file_location("inscripcion_nadadores", "1-inscripcion_nadadores.py")
inscripcion_nadadores = importlib.util.module_from_spec(spec)
spec.loader.exec_module(inscripcion_nadadores)

# Importar los otros módulos con sus nuevos nombres
spec1 = importlib.util.spec_from_file_location("generar_sembrado", "2-generar_sembrado.py")
script1 = importlib.util.module_from_spec(spec1)
spec1.loader.exec_module(script1)

spec2 = importlib.util.spec_from_file_location("generar_sembrado_por_tiempo", "3-generar_sembrado_por_tiempo.py")
script2 = importlib.util.module_from_spec(spec2)
spec2.loader.exec_module(script2)

spec3 = importlib.util.spec_from_file_location("procesar_resultados", "4-procesar_resultados.py")
script3 = importlib.util.module_from_spec(spec3)
spec3.loader.exec_module(script3)

# Importar ambos módulos de papeletas
spec4 = importlib.util.spec_from_file_location("generar_papeletas", "generar_papeletas.py")
papeletas_pdf_module = importlib.util.module_from_spec(spec4)
spec4.loader.exec_module(papeletas_pdf_module)

spec5 = importlib.util.spec_from_file_location("generar_papeletas_excel", "generar_papeletas_excel.py")
papeletas_excel_module = importlib.util.module_from_spec(spec5)
spec5.loader.exec_module(papeletas_excel_module)

# Importar el gestor de eventos
spec6 = importlib.util.spec_from_file_location("event_manager", "event_manager.py")
event_manager_module = importlib.util.module_from_spec(spec6)
spec6.loader.exec_module(event_manager_module)

#Configuración de la página
st.set_page_config(
    page_title="TEN - Gestión de Competencias",
    page_icon="🏊‍♀️",
    layout="wide",
    initial_sidebar_state="expanded"
)

#CSS personalizado con colores del logo
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1E88E5 0%, #64B5F6 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
    }
    
    .main-header h1 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: bold;
    }
    
    .main-header p {
        margin: 0.5rem 0 0 0;
        font-size: 1.2rem;
        opacity: 0.9;
    }
    
    .stButton > button {
        background: linear-gradient(90deg, #1E88E5 0%, #64B5F6 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 2rem;
        font-weight: bold;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        background: linear-gradient(90deg, #1565C0 0%, #42A5F5 100%);
        box-shadow: 0 4px 15px rgba(30, 136, 229, 0.3);
        transform: translateY(-2px);
    }
    
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 4px solid #1E88E5;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }
    
    .feature-card h3 {
        color: #1E88E5;
        margin-top: 0;
    }
    
    .sidebar .stSelectbox label {
        color: #1E88E5;
        font-weight: bold;
    }
    
    .success-message {
        background: #E8F5E8;
        color: #2E7D2E;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #4CAF50;
        margin: 1rem 0;
    }
    
    .info-message {
        background: #E3F2FD;
        color: #1565C0;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #1E88E5;
        margin: 1rem 0;
    }
    
    .warning-message {
        background: #FFF3E0;
        color: #E65100;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #FF9800;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

#def load_script_functions(): # No longer needed
#    """Cargar funciones de los scripts existentes"""
#    functions = {}
#    
#    # Cargar funciones del script 1
#    spec1 = importlib.util.spec_from_file_location("script1", "1-generar_sembrado.py")
#    script1 = importlib.util.module_from_spec(spec1)
#    spec1.loader.exec_module(script1)
#    functions['generar_sembrado_categoria'] = script1.main
#    
#    # Cargar funciones del script 2
#    spec2 = importlib.util.spec_from_file_location("script2", "2-generar_sembrado_por_tiempo.py")
#    script2 = importlib.util.module_from_spec(spec2)
#    spec2.loader.exec_module(script2)
#    functions['generar_sembrado_tiempo'] = script2.main
#    
#    # Cargar funciones del script 3
#    spec3 = importlib.util.spec_from_file_location("script3", "3-procesar_resultados.py")
#    script3 = importlib.util.module_from_spec(spec3)
#    spec3.loader.exec_module(script3)
#    functions['procesar_resultados'] = script3.main_full
#    
#    return functions

def main():
    # Header principal
    st.markdown("""
    <div class="main-header">
        <h1>🏊‍♀️ TEN - Gestión de Competencias</h1>
        <p>Sistema completo para administrar competencias de natación</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    st.sidebar.image("img/TEN.png", width=150)
    st.sidebar.markdown("## 📋 Panel de Control")
    
    # Selector de funciones
    default_index = 0
    if 'selected_option' in st.session_state:
        options = [
            "🏠 Inicio",
            "🎯 Creación del Evento",
            "✍️ Inscripción de Nadadores",
            "📊 Sembrado de Competencia",
            "📋 Generar Papeletas",
            "🏆 Procesar Resultados",
            "📁 Gestión de Archivos"
        ]
        if st.session_state.selected_option in options:
            default_index = options.index(st.session_state.selected_option)

    opcion = st.sidebar.selectbox(
        "Selecciona una operación:",
        [
            "🏠 Inicio",
            "🎯 Creación del Evento",
            "✍️ Inscripción de Nadadores",
            "📊 Sembrado de Competencia",
            "📋 Generar Papeletas",
            "🏆 Procesar Resultados",
            "📁 Gestión de Archivos"
        ],
        index=default_index
    )
    
    if opcion == "🏠 Inicio":
        mostrar_inicio()
    elif opcion == "🎯 Creación del Evento":
        mostrar_creacion_evento()
    elif opcion == "✍️ Inscripción de Nadadores":
        inscripcion_nadadores_interface()
    elif opcion == "📊 Sembrado de Competencia":
        sembrado_competencia_interface()
    elif opcion == "📋 Generar Papeletas":
        generar_papeletas_interface()
    elif opcion == "🏆 Procesar Resultados":
        procesar_resultados()
    elif opcion == "📁 Gestión de Archivos":
        gestion_archivos()

def mostrar_inicio():
    st.markdown("## 🏊‍♀️ Bienvenido al Sistema TEN")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <h3>🎯 Creación del Evento</h3>
            <p>Configure el nombre, pruebas disponibles y rango de edades para su competencia de natación.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="feature-card">
            <h3>📊 Sembrado por Categoría</h3>
            <p>Organiza las series agrupando nadadores por categoría de edad y luego por tiempo dentro de cada categoría.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="feature-card">
            <h3>🏆 Procesamiento de Resultados</h3>
            <p>Genera reportes de premiación con sistema de puntos y clasificaciones por categoría y equipos.</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <h3>✍️ Inscripción de Nadadores</h3>
            <p>Registra nuevos nadadores con sus datos personales y tiempos de inscripción por prueba.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="feature-card">
            <h3>⏱️ Sembrado por Tiempo</h3>
            <p>Crea series basándose únicamente en los tiempos de inscripción, sin importar la categoría.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="feature-card">
            <h3>📁 Gestión de Archivos</h3>
            <p>Sube y descarga archivos Excel, visualiza datos y administra los archivos del sistema.</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-message">
        <h4>📋 Flujo de trabajo recomendado:</h4>
        <ol>
            <li><strong>Crea el evento</strong> definiendo nombre, pruebas y rango de edades</li>
            <li><strong>Inscribe nadadores</strong> usando el formulario de inscripción integrado</li>
            <li>Genera el sembrado (por categoría o tiempo)</li>
            <li>Después de la competencia, procesa los resultados</li>
            <li>Descarga los reportes generados</li>
        </ol>
        <p><em>Alternativamente, puedes subir un archivo <strong>planilla_inscripcion.xlsx</strong> existente en "Gestión de Archivos"</em></p>
    </div>
    """, unsafe_allow_html=True)

def mostrar_creacion_evento():
    st.markdown("## 🎯 Creación del Evento")

    event_manager = event_manager_module.EventManager()

    # Verificar si ya existe un evento configurado
    event_info = event_manager.get_event_info()

    if event_info:
        # Validar configuración completa
        is_complete, validation_message = event_manager.validate_event_configuration()

        if is_complete:
            st.markdown(f"""
            <div class="success-message">
                <h4>✅ Evento configurado: {event_info['name']}</h4>
                <p><strong>Categorías:</strong> {len(event_info['categories'])}</p>
                <p><strong>Pruebas del evento:</strong> {len(event_info['events'])}</p>
                <p><strong>Rango de edades:</strong> {event_info['min_age']} - {event_info['max_age']} años</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-message">
                <h4>⚠️ Evento configurado: {event_info['name']}</h4>
                <p><strong>Estado:</strong> Configuración incompleta</p>
                <p><strong>Errores:</strong> {validation_message}</p>
            </div>
            """, unsafe_allow_html=True)

        # Mostrar detalles del evento en expandibles
        with st.expander("Ver detalles del evento"):
            # Mostrar mensaje de bienvenida si existe
            welcome_message = event_info.get('welcome_message', '').strip()
            if welcome_message:
                st.markdown("**📝 Mensaje de Bienvenida:**")
                st.text_area("", value=welcome_message, height=100, disabled=True, key="detail_welcome_message")

            # Mostrar mensaje de despedida si existe
            farewell_message = event_info.get('farewell_message', '').strip()
            if farewell_message:
                st.markdown("**📝 Mensaje de Despedida:**")
                st.text_area("", value=farewell_message, height=100, disabled=True, key="detail_farewell_message")

            if welcome_message or farewell_message:
                st.markdown("---")

            col1, col2 = st.columns(2)

            with col1:
                st.write("**Categorías:**")
                for i, category in enumerate(event_info['categories'], 1):
                    age_info = f" ({category['age_range']})" if category['age_range'] else ""
                    st.write(f"{i}. {category['name']}{age_info}")

            with col2:
                st.write("**Pruebas del evento (orden):**")
                for i, event in enumerate(event_info['events'], 1):
                    st.write(f"{i}. {event}")

            # Mostrar asignación de pruebas por categoría
            if event_info['category_events']:
                st.write("**Asignación de pruebas por categoría:**")
                for cat_name, events in event_info['category_events'].items():
                    st.write(f"**{cat_name}:** {', '.join(events) if events else 'Sin pruebas asignadas'}")

        # Opciones para modificar o eliminar
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            if st.button("🔄 Modificar Evento", type="secondary"):
                st.session_state.modificar_evento = True
                st.rerun()

        with col2:
            if st.button("📄 Generar PDF", type="secondary"):
                with st.spinner("Generando reporte PDF..."):
                    try:
                        result = event_manager.generate_event_pdf_report()
                        if result is None:
                            st.error("❌ La función generate_event_pdf_report devolvió None")
                        else:
                            pdf_data, filename = result
                            if pdf_data:
                                st.download_button(
                                    label="📥 Descargar Reporte PDF",
                                    data=pdf_data,
                                    file_name=filename,
                                    mime="application/pdf"
                                )
                                st.success("✅ Reporte PDF generado exitosamente")
                            else:
                                st.error(f"❌ Error generando PDF: {filename}")
                    except Exception as e:
                        st.error(f"❌ Error inesperado generando PDF: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())

        with col3:
            if st.button("🗑️ Eliminar Evento", type="secondary"):
                success, message = event_manager.delete_event_config()
                if success:
                    st.success(message)
                    st.rerun()
                else:
                    st.error(message)

        with col4:
            if st.button("➡️ Ir a Inscripción", type="primary", disabled=not is_complete):
                if is_complete:
                    st.session_state.selected_option = "✍️ Inscripción de Nadadores"
                    st.rerun()

    # Formulario para crear/modificar evento
    if not event_info or st.session_state.get('modificar_evento', False):
        crear_formulario_evento(event_manager, event_info)


def crear_formulario_evento(event_manager, event_info=None):
    """Crear formulario completo para configuración de evento"""

    if event_info:
        st.markdown("### 🔄 Modificar Evento Existente")
    else:
        st.markdown("### ➕ Crear Nuevo Evento")
        st.markdown("""
        <div class="info-message">
            Configure los detalles completos de su evento de natación siguiendo el orden:
            <strong>1) Nombre y edades → 2) Categorías → 3) Pruebas del evento → 4) Asignación por categoría</strong>
        </div>
        """, unsafe_allow_html=True)

    # Inicializar session_state si es necesario
    if 'evento_step' not in st.session_state:
        st.session_state.evento_step = 1

    if 'evento_categories' not in st.session_state:
        st.session_state.evento_categories = event_info['categories'] if event_info else []

    if 'evento_event_order' not in st.session_state:
        st.session_state.evento_event_order = event_info['events'] if event_info else []

    if 'evento_category_events' not in st.session_state:
        st.session_state.evento_category_events = event_info['category_events'] if event_info else {}

    # Pasos del formulario
    tabs = st.tabs(["1️⃣ Datos Básicos", "2️⃣ Categorías", "3️⃣ Pruebas del Evento", "4️⃣ Asignación por Categoría", "5️⃣ Finalizar"])

    with tabs[0]:
        mostrar_paso_datos_basicos(event_manager, event_info)

    with tabs[1]:
        mostrar_paso_categorias(event_manager)

    with tabs[2]:
        mostrar_paso_pruebas_evento(event_manager)

    with tabs[3]:
        mostrar_paso_asignacion_categorias(event_manager)

    with tabs[4]:
        mostrar_paso_finalizar(event_manager, event_info)


def mostrar_paso_datos_basicos(event_manager, event_info):
    """Paso 1: Datos básicos del evento"""
    st.markdown("### 📝 Información Básica del Evento")

    # Nombre del evento
    event_name = st.text_input(
        "Nombre del Evento",
        value=event_info['name'] if event_info else "",
        placeholder="Ej: Campeonato Nacional de Natación 2024",
        key="evento_name"
    )

    # Fechas del evento
    st.markdown("**📅 Fechas del Evento:**")
    col1, col2 = st.columns(2)

    # Parsear fechas existentes si están disponibles
    from datetime import date
    default_start_date = date.today()
    default_end_date = date.today()

    if event_info:
        try:
            if event_info.get('start_date'):
                default_start_date = datetime.fromisoformat(event_info['start_date']).date()
            if event_info.get('end_date'):
                default_end_date = datetime.fromisoformat(event_info['end_date']).date()
        except:
            pass

    with col1:
        start_date = st.date_input(
            "Fecha de Inicio",
            value=default_start_date,
            key="evento_start_date",
            help="Fecha de inicio del evento"
        )

    with col2:
        end_date = st.date_input(
            "Fecha de Finalización",
            value=default_end_date,
            key="evento_end_date",
            help="Fecha de finalización del evento"
        )

    # Rango de edades
    st.markdown("**🎂 Rango de Edades Permitido:**")
    col1, col2 = st.columns(2)
    with col1:
        min_age = st.number_input(
            "Edad Mínima",
            min_value=5,
            max_value=80,
            value=event_info['min_age'] if event_info else 8,
            step=1,
            key="evento_min_age"
        )

    with col2:
        max_age = st.number_input(
            "Edad Máxima",
            min_value=5,
            max_value=80,
            value=event_info['max_age'] if event_info else 18,
            step=1,
            key="evento_max_age"
        )

    # Criterio de edad
    st.markdown("**📅 Criterio para Determinar la Edad:**")
    age_criteria_options = {
        "event_date": "Edad el día del evento",
        "december_31": "Edad al 31 de diciembre"
    }

    default_criteria = event_info.get('age_criteria', 'event_date') if event_info else 'event_date'
    age_criteria = st.selectbox(
        "Seleccione cómo se calculará la edad de los nadadores",
        options=list(age_criteria_options.keys()),
        format_func=lambda x: age_criteria_options[x],
        index=list(age_criteria_options.keys()).index(default_criteria),
        key="evento_age_criteria",
        help="Defina si la edad se calcula al día del evento o al 31 de diciembre del año en curso"
    )

    # Valores de inscripción
    st.markdown("**💰 Valores de Inscripción:**")
    col1, col2 = st.columns(2)
    with col1:
        swimmer_fee = st.number_input(
            "Valor por Nadador ($)",
            min_value=0,
            value=event_info.get('swimmer_fee', 0) if event_info else 0,
            step=1000,
            key="evento_swimmer_fee",
            help="Costo de inscripción por cada nadador"
        )

    with col2:
        team_fee = st.number_input(
            "Valor por Equipo ($)",
            min_value=0,
            value=event_info.get('team_fee', 0) if event_info else 0,
            step=5000,
            key="evento_team_fee",
            help="Costo de inscripción por equipo/club"
        )

    # Mensaje de bienvenida
    st.markdown("**📝 Mensaje de Bienvenida (Opcional):**")
    welcome_message = st.text_area(
        "Mensaje de bienvenida para el evento",
        value=event_info.get('welcome_message', '') if event_info else '',
        max_chars=1000,
        height=100,
        key="evento_welcome_message",
        help="Mensaje que aparecerá en los reportes y documentos del evento (máximo 1000 caracteres)",
        placeholder="Escriba aquí un mensaje de bienvenida para los participantes del evento..."
    )

    # Mensaje de despedida
    st.markdown("**📝 Mensaje de Despedida (Opcional):**")
    farewell_message = st.text_area(
        "Mensaje de despedida para el evento",
        value=event_info.get('farewell_message', '') if event_info else '',
        max_chars=1000,
        height=100,
        key="evento_farewell_message",
        help="Mensaje de cierre que aparecerá en los reportes y documentos del evento (máximo 1000 caracteres)",
        placeholder="Escriba aquí un mensaje de despedida para los participantes del evento..."
    )

    # Logo del evento
    st.markdown("**🖼️ Logo del Evento (Opcional):**")
    uploaded_logo = st.file_uploader(
        "Subir logo del evento",
        type=['png', 'jpg', 'jpeg', 'gif'],
        key="evento_logo_upload",
        help="Imagen que aparecerá en los reportes y documentos del evento"
    )

    # Mostrar logo actual si existe
    if event_info and event_info.get('event_logo'):
        logo_path = f"event_logos/{event_info['event_logo']}"
        if os.path.exists(logo_path):
            st.image(logo_path, caption="Logo actual del evento", width=200)

    # Validaciones básicas
    errors = []
    if event_name and len(event_name.strip()) < 3:
        errors.append("El nombre del evento debe tener al menos 3 caracteres")
    if min_age >= max_age:
        errors.append("La edad mínima debe ser menor que la máxima")

    if errors:
        for error in errors:
            st.error(f"❌ {error}")
    elif event_name and event_name.strip() and min_age < max_age:
        st.success("✅ Datos básicos completados correctamente")

    # Información del mensaje de bienvenida
    if welcome_message:
        char_count = len(welcome_message)
        if char_count > 1000:
            st.warning(f"⚠️ El mensaje excede el límite de 1000 caracteres ({char_count}/1000)")
        else:
            st.info(f"📊 Caracteres utilizados: {char_count}/1000")


def mostrar_paso_categorias(event_manager):
    """Paso 2: Gestión de categorías"""
    st.markdown("### 🏷️ Configuración de Categorías")

    # Pestañas para creación manual vs carga desde Excel
    tab1, tab2 = st.tabs(["Creación Manual", "Cargar desde Excel"])

    with tab1:
        mostrar_creacion_manual_categorias()

    with tab2:
        mostrar_carga_excel_categorias(event_manager)

    # Mostrar categorías actuales
    if st.session_state.evento_categories:
        st.markdown("### 📋 Categorías Configuradas")
        mostrar_lista_categorias_editable()


def mostrar_creacion_manual_categorias():
    """Interfaz para creación manual de categorías"""
    st.markdown("**Agregar nueva categoría:**")

    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])

    with col1:
        new_category_name = st.text_input("Nombre de la categoría", placeholder="Ej: Juvenil A", key="new_cat_name")

    with col2:
        min_age_cat = st.number_input("Edad inicial", min_value=5, max_value=80, value=12, key="new_cat_min_age")

    with col3:
        max_age_cat = st.number_input("Edad final", min_value=5, max_value=80, value=13, key="new_cat_max_age")

    with col4:
        if st.button("➕ Agregar", type="primary", key="add_category_btn"):
            if new_category_name.strip():
                # Validar rango de edad
                if min_age_cat > max_age_cat:
                    st.error("La edad inicial debe ser menor o igual que la final")
                else:
                    # Crear rango de edad formateado
                    age_range = event_manager_module.EventManager().format_age_range(min_age_cat, max_age_cat)

                    # Validar nombre único
                    is_valid, message = event_manager_module.EventManager().validate_category_name(
                        new_category_name.strip(),
                        st.session_state.evento_categories
                    )

                    if is_valid:
                        new_category = {
                            'name': new_category_name.strip(),
                            'age_range': age_range
                        }
                        st.session_state.evento_categories.append(new_category)
                        # Ordenar categorías por edad
                        st.session_state.evento_categories = event_manager_module.EventManager().sort_categories_by_age(
                            st.session_state.evento_categories)
                        # Limpiar campos inmediatamente
                        st.session_state.new_cat_name = ""
                        st.success(f"✅ Categoría '{new_category_name.strip()}' agregada exitosamente")
                        st.rerun()
                    else:
                        st.error(message)
            else:
                st.error("El nombre de la categoría es requerido")


def mostrar_carga_excel_categorias(event_manager):
    """Interfaz para cargar categorías desde Excel"""
    st.markdown("""
    **Cargar desde archivo Excel:**

    El archivo debe contener columnas con nombres que incluyan:
    - **Nombre/Categoría:** nombre de la categoría
    - **Edad/Rango:** rango de edades (opcional)
    """)

    uploaded_file = st.file_uploader(
        "Seleccionar archivo Excel",
        type=['xlsx', 'xls'],
        key="categories_upload"
    )

    if uploaded_file:
        if st.button("📤 Cargar Categorías", type="primary", key="load_categories_btn"):
            success, result = event_manager.load_categories_from_excel(uploaded_file)

            if success:
                # Sobrescribir categorías existentes
                st.session_state.evento_categories = result
                st.success(f"✅ {len(result)} categorías cargadas exitosamente")
                st.rerun()
            else:
                st.error(f"❌ Error al cargar categorías: {result}")


def mostrar_lista_categorias_editable():
    """Mostrar lista de categorías con opciones de edición"""
    # Ordenar categorías antes de mostrar
    st.session_state.evento_categories = event_manager_module.EventManager().sort_categories_by_age(
        st.session_state.evento_categories)

    for i, category in enumerate(st.session_state.evento_categories):
        col1, col2, col3, col4, col5 = st.columns([2, 1, 1, 1, 1])

        with col1:
            # Campo editable para nombre
            new_name = st.text_input(
                f"Categoría {i+1}",
                value=category['name'],
                key=f"cat_name_{i}",
                label_visibility="collapsed"
            )

        # Parsear el rango de edad actual
        event_mgr = event_manager_module.EventManager()
        min_age, max_age = event_mgr.parse_age_range(category.get('age_range', '12-13'))
        if min_age is None or max_age is None:
            min_age, max_age = 12, 13

        with col2:
            # Campo para edad inicial
            new_min_age = st.number_input(
                f"Min {i+1}",
                min_value=5,
                max_value=80,
                value=min_age,
                key=f"cat_min_age_{i}",
                label_visibility="collapsed"
            )

        with col3:
            # Campo para edad final
            new_max_age = st.number_input(
                f"Max {i+1}",
                min_value=5,
                max_value=80,
                value=max_age,
                key=f"cat_max_age_{i}",
                label_visibility="collapsed"
            )

        with col4:
            # Botón actualizar
            if st.button("💾", key=f"update_cat_{i}", help="Actualizar categoría"):
                if new_name.strip():
                    if new_min_age > new_max_age:
                        st.error("La edad inicial debe ser menor o igual que la final")
                    else:
                        # Validar nombre único (excluyendo la actual)
                        is_valid, message = event_mgr.validate_category_name(
                            new_name.strip(),
                            st.session_state.evento_categories,
                            exclude_index=i
                        )

                        if is_valid:
                            new_age_range = event_mgr.format_age_range(new_min_age, new_max_age)
                            st.session_state.evento_categories[i] = {
                                'name': new_name.strip(),
                                'age_range': new_age_range
                            }
                            # Reordenar después de la actualización
                            st.session_state.evento_categories = event_mgr.sort_categories_by_age(
                                st.session_state.evento_categories)
                            st.success(f"Categoría {i+1} actualizada")
                            st.rerun()
                        else:
                            st.error(message)
                else:
                    st.error("El nombre no puede estar vacío")

        with col5:
            # Botón eliminar
            if st.button("🗑️", key=f"delete_cat_{i}", help="Eliminar categoría"):
                st.session_state.evento_categories.pop(i)
                st.success("Categoría eliminada")
                st.rerun()


def mostrar_paso_pruebas_evento(event_manager):
    """Paso 3: Selección y orden de pruebas del evento"""
    st.markdown("### 🏊‍♀️ Configuración de Pruebas del Evento")

    st.markdown("""
    <div class="info-message">
        <strong>Instrucciones:</strong> Seleccione las pruebas disponibles de la lista de la izquierda
        y arrástrelas a la lista de la derecha para definir el orden del evento.
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([2, 1, 2])

    # Lista de pruebas disponibles
    with col1:
        st.markdown("**🏊 Pruebas Disponibles**")
        all_events = event_manager.get_available_events()
        selected_events = st.session_state.evento_event_order

        # Filtrar eventos que ya no están en el orden
        available_events = [event for event in all_events if event not in selected_events]

        if available_events:
            for event in available_events:
                col_a, col_b = st.columns([3, 1])
                with col_a:
                    # Obtener restricción de edad para el evento
                    min_age = event_manager.get_event_age_restriction(event)
                    st.write(f"• {event} <small>(≥{min_age} años)</small>", unsafe_allow_html=True)
                with col_b:
                    if st.button("➡️", key=f"add_event_{event}", help=f"Agregar {event} (edad mínima: {min_age} años)"):
                        st.session_state.evento_event_order.append(event)
                        st.rerun()
        else:
            st.info("Todas las pruebas han sido seleccionadas")

    # Separador visual
    with col2:
        st.markdown("<br>" * 8, unsafe_allow_html=True)
        st.markdown("**⬅️ ➡️**", unsafe_allow_html=True)

    # Lista de pruebas seleccionadas (orden del evento)
    with col3:
        st.markdown("**🏆 Pruebas del Evento (Orden)**")

        if st.session_state.evento_event_order:
            for i, event in enumerate(st.session_state.evento_event_order):
                col_a, col_b, col_c, col_d = st.columns([1, 3, 1, 1])

                with col_a:
                    st.write(f"{i+1}.")

                with col_b:
                    st.write(event)

                with col_c:
                    # Botones para reordenar
                    if i > 0 and st.button("⬆️", key=f"up_{i}", help="Subir"):
                        # Intercambiar con el anterior
                        st.session_state.evento_event_order[i], st.session_state.evento_event_order[i-1] = \
                        st.session_state.evento_event_order[i-1], st.session_state.evento_event_order[i]
                        st.rerun()

                with col_d:
                    # Botón para remover
                    if st.button("❌", key=f"remove_event_{i}", help="Quitar del evento"):
                        st.session_state.evento_event_order.pop(i)
                        st.rerun()

        else:
            st.info("Agregue pruebas desde la lista de la izquierda")

        # Botones adicionales
        if st.session_state.evento_event_order:
            if st.button("🔄 Limpiar Orden"):
                st.session_state.evento_event_order = []
                st.rerun()


def mostrar_paso_asignacion_categorias(event_manager):
    """Paso 4: Asignación de pruebas por categoría"""
    st.markdown("### 🎯 Asignación de Pruebas por Categoría")

    if not st.session_state.evento_categories:
        st.warning("⚠️ Primero debe configurar las categorías en el paso anterior")
        return

    if not st.session_state.evento_event_order:
        st.warning("⚠️ Primero debe configurar las pruebas del evento en el paso anterior")
        return

    st.markdown("""
    <div class="info-message">
        Seleccione qué pruebas puede nadar cada categoría. Solo aparecen las pruebas
        que fueron incluidas en el evento.
    </div>
    """, unsafe_allow_html=True)

    # Para cada categoría, mostrar checkboxes de las pruebas del evento
    for category in st.session_state.evento_categories:
        category_name = category['name']

        st.markdown(f"**📋 {category_name}** {f'({category["age_range"]})' if category['age_range'] else ''}")

        # Obtener pruebas actualmente asignadas a esta categoría
        current_events = st.session_state.evento_category_events.get(category_name, [])

        # Crear checkboxes en columnas
        cols = st.columns(3)
        selected_events_for_category = []

        for i, event in enumerate(st.session_state.evento_event_order):
            col = cols[i % 3]
            with col:
                is_selected = st.checkbox(
                    event,
                    value=event in current_events,
                    key=f"cat_event_{category_name}_{i}"
                )
                if is_selected:
                    selected_events_for_category.append(event)

        # Actualizar session state
        st.session_state.evento_category_events[category_name] = selected_events_for_category

        # Mostrar resumen
        if selected_events_for_category:
            st.success(f"✅ {len(selected_events_for_category)} pruebas seleccionadas para {category_name}")
        else:
            st.error(f"❌ No hay pruebas seleccionadas para {category_name}")

        st.divider()


def mostrar_paso_finalizar(event_manager, event_info):
    """Paso 5: Finalizar y guardar configuración"""
    st.markdown("### ✅ Finalizar Configuración del Evento")

    # Validaciones finales
    errors = []
    warnings = []

    # Validar nombre del evento
    event_name = st.session_state.get('evento_name', '').strip()
    if not event_name or len(event_name) < 3:
        errors.append("El nombre del evento debe tener al menos 3 caracteres")

    # Validar edades
    min_age = st.session_state.get('evento_min_age', 0)
    max_age = st.session_state.get('evento_max_age', 0)
    if min_age >= max_age:
        errors.append("El rango de edades no es válido")

    # Validar fechas
    start_date = st.session_state.get('evento_start_date')
    end_date = st.session_state.get('evento_end_date')
    if start_date and end_date and start_date > end_date:
        errors.append("La fecha de inicio no puede ser posterior a la fecha de finalización")

    # Validar categorías
    if not st.session_state.evento_categories:
        errors.append("Debe configurar al menos una categoría")

    # Validar pruebas del evento
    if not st.session_state.evento_event_order:
        errors.append("Debe seleccionar al menos una prueba para el evento")

    # Validar asignación de pruebas por categoría
    for category in st.session_state.evento_categories:
        cat_name = category['name']
        cat_events = st.session_state.evento_category_events.get(cat_name, [])
        if not cat_events:
            errors.append(f"La categoría '{cat_name}' no tiene pruebas asignadas")

    # Mostrar resumen
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**📋 Resumen de Configuración:**")
        st.write(f"**Nombre:** {event_name}")

        # Mostrar fechas si están disponibles
        if start_date and end_date:
            st.write(f"**Fechas:** {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
        elif start_date:
            st.write(f"**Fecha de inicio:** {start_date.strftime('%d/%m/%Y')}")
        elif end_date:
            st.write(f"**Fecha de finalización:** {end_date.strftime('%d/%m/%Y')}")

        st.write(f"**Edades:** {min_age} - {max_age} años")

        # Mostrar criterio de edad
        age_criteria = st.session_state.get('evento_age_criteria', 'event_date')
        age_criteria_text = "Edad al 31 de diciembre" if age_criteria == 'december_31' else "Edad el día del evento"
        st.write(f"**Criterio de edad:** {age_criteria_text}")

        # Mostrar valores de inscripción
        swimmer_fee = st.session_state.get('evento_swimmer_fee', 0)
        team_fee = st.session_state.get('evento_team_fee', 0)
        if swimmer_fee > 0:
            st.write(f"**Valor por nadador:** ${swimmer_fee:,}")
        if team_fee > 0:
            st.write(f"**Valor por equipo:** ${team_fee:,}")

        st.write(f"**Categorías:** {len(st.session_state.evento_categories)}")
        st.write(f"**Pruebas del evento:** {len(st.session_state.evento_event_order)}")

        # Mostrar mensaje de bienvenida si existe
        welcome_message = st.session_state.get('evento_welcome_message', '').strip()
        if welcome_message:
            st.write("**📝 Mensaje de bienvenida:**")
            st.text_area("", value=welcome_message, height=100, disabled=True, key="preview_welcome_message")

        # Mostrar mensaje de despedida si existe
        farewell_message = st.session_state.get('evento_farewell_message', '').strip()
        if farewell_message:
            st.write("**📝 Mensaje de despedida:**")
            st.text_area("", value=farewell_message, height=100, disabled=True, key="preview_farewell_message")

    with col2:
        st.markdown("**🎯 Categorías y sus pruebas:**")
        for category in st.session_state.evento_categories:
            cat_name = category['name']
            cat_events = st.session_state.evento_category_events.get(cat_name, [])
            st.write(f"**{cat_name}:** {len(cat_events)} pruebas")

    # Mostrar errores o warnings
    if errors:
        st.error("❌ **Errores que deben corregirse:**\n" + "\n".join([f"• {error}" for error in errors]))
    elif warnings:
        st.warning("⚠️ **Advertencias:**\n" + "\n".join([f"• {warning}" for warning in warnings]))
    else:
        st.success("✅ **Configuración completa y válida**")

    # Botones finales
    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        if st.button("💾 Guardar Evento", type="primary", disabled=bool(errors)):
            # Obtener valores adicionales
            swimmer_fee = st.session_state.get('evento_swimmer_fee', 0)
            team_fee = st.session_state.get('evento_team_fee', 0)
            welcome_message = st.session_state.get('evento_welcome_message', '')
            uploaded_logo = st.session_state.get('evento_logo_upload')
            start_date = st.session_state.get('evento_start_date')
            end_date = st.session_state.get('evento_end_date')
            age_criteria = st.session_state.get('evento_age_criteria', 'event_date')

            # Procesar logo si se subió uno nuevo
            event_logo = None
            if uploaded_logo and event_name:
                logo_success, logo_result = event_manager.save_event_logo(uploaded_logo, event_name)
                if logo_success:
                    event_logo = logo_result
                    st.info(f"✅ Logo guardado: {logo_result}")
                else:
                    st.warning(f"⚠️ Error guardando logo: {logo_result}")
            elif event_info and event_info.get('event_logo'):
                # Mantener logo existente si no se subió uno nuevo
                event_logo = event_info['event_logo']

            success, message = event_manager.save_event_config(
                event_name,
                st.session_state.evento_categories,
                st.session_state.evento_event_order,
                st.session_state.evento_category_events,
                min_age,
                max_age,
                swimmer_fee,
                team_fee,
                welcome_message,
                farewell_message,
                event_logo,
                start_date,
                end_date,
                age_criteria
            )

            if success:
                st.success(message)
                # Limpiar session state
                for key in list(st.session_state.keys()):
                    if key.startswith('evento_'):
                        del st.session_state[key]
                st.session_state.modificar_evento = False
                st.rerun()
            else:
                st.error(message)

    with col2:
        if st.button("🔄 Reiniciar Formulario"):
            # Limpiar session state
            for key in list(st.session_state.keys()):
                if key.startswith('evento_'):
                    del st.session_state[key]
            st.rerun()

    with col3:
        if st.button("❌ Cancelar"):
            # Limpiar session state
            for key in list(st.session_state.keys()):
                if key.startswith('evento_'):
                    del st.session_state[key]
            st.session_state.modificar_evento = False
            st.rerun()


def generar_sembrado_categoria():
    st.markdown("## 📊 Generar Sembrado por Categoría")
    
    st.markdown("""
    <div class="info-message">
        Este proceso agrupa los nadadores por categoría de edad y luego los ordena por tiempo dentro de cada categoría.
        Las series se organizan con los nadadores más rápidos en las últimas series.
    </div>
    """, unsafe_allow_html=True)
    
    # Verificar archivo de entrada
    if not os.path.exists("planilla_inscripcion.xlsx"):
        st.markdown("""
        <div class="warning-message">
            ⚠️ No se encontró el archivo <strong>planilla_inscripcion.xlsx</strong>. 
            Por favor, súbelo en la sección "Gestión de Archivos".
        </div>
        """, unsafe_allow_html=True)
        return
    
    # Mostrar información del archivo
    try:
        df = pd.read_excel("planilla_inscripcion.xlsx")
        st.success(f"✅ Archivo cargado: {len(df)} nadadores registrados")
        
        # Mostrar preview
        if st.checkbox("Ver vista previa de datos"):
            st.dataframe(df.head(10))
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return
    
    col1, col2 = st.columns([1, 3])
    
    with col1:
        if st.button("🚀 Generar Sembrado por Categoría", type="primary"):
            with st.spinner("Generando sembrado..."):
                try:
                    # Ejecutar el script
                    #result = subprocess.run([sys.executable, "1-generar_sembrado.py"],
                    #                      capture_output=True, text=True)
                    script1.main() # Llamar la función directamente
                    
                    #if result.returncode == 0:
                    st.markdown("""
                        <div class="success-message">
                            ✅ <strong>Sembrado generado exitosamente!</strong><br>
                            Archivo creado: <code>sembrado_competencia.xlsx</code>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Mostrar output
                    #if result.stdout:
                    #    st.text("Output:")
                    #    st.text(result.stdout)
                    
                    #else:
                    #    st.error(f"Error al generar sembrado: {result.stderr}")
                        
                except Exception as e:
                    st.error(f"Error al ejecutar el script: {e}")
    
    with col2:
        if os.path.exists("sembrado_competencia.xlsx"):
            st.info("📄 Archivo generado disponible para descarga")
            
            # Botón de descarga
            with open("sembrado_competencia.xlsx", "rb") as file:
                st.download_button(
                    label="⬇️ Descargar Sembrado por Categoría",
                    data=file.read(),
                    file_name="sembrado_competencia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Sección de papeletas para jueces
            st.markdown("### 📋 Generar Papeletas para Jueces")
            st.info("Las papeletas se generan en formato Excel con rectángulos imprimibles")
            
            if st.button("📄 Generar Papeletas Excel", type="secondary", key="papeletas_excel_cat"):
                with st.spinner("Generando papeletas en Excel..."):
                    success, message = papeletas_excel_module.generar_papeletas_excel()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
            
            # Botón de descarga para papeletas Excel si existe
            if os.path.exists("papeletas_jueces.xlsx"):
                st.info("📄 Papeletas Excel disponibles")
                with open("papeletas_jueces.xlsx", "rb") as file:
                    st.download_button(
                        label="⬇️ Descargar Papeletas Excel",
                        data=file.read(),
                        file_name="papeletas_jueces.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

def generar_sembrado_tiempo():
    st.markdown("## ⏱️ Generar Sembrado por Tiempo")
    
    st.markdown("""
    <div class="info-message">
        Este proceso ordena todos los nadadores únicamente por tiempo de inscripción, 
        sin importar la categoría de edad. Ideal para competencias open o clasificatorias.
    </div>
    """, unsafe_allow_html=True)
    
    if not os.path.exists("planilla_inscripcion.xlsx"):
        st.markdown("""
        <div class="warning-message">
            ⚠️ No se encontró el archivo <strong>planilla_inscripcion.xlsx</strong>. 
            Por favor, súbelo en la sección "Gestión de Archivos".
        </div>
        """, unsafe_allow_html=True)
        return
    
    try:
        df = pd.read_excel("planilla_inscripcion.xlsx")
        st.success(f"✅ Archivo cargado: {len(df)} nadadores registrados")
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return
    
    col1, col2 = st.columns([1, 3])
    
    with col1:
        if st.button("🚀 Generar Sembrado por Tiempo", type="primary"):
            with st.spinner("Generando sembrado..."):
                try:
                    #result = subprocess.run([sys.executable, "2-generar_sembrado_por_tiempo.py"],
                    #                      capture_output=True, text=True)
                    script2.main() # Llamar la función directamente
                    
                    #if result.returncode == 0:
                    st.markdown("""
                        <div class="success-message">
                            ✅ <strong>Sembrado por tiempo generado exitosamente!</strong><br>
                            Archivo creado: <code>sembrado_competencia_POR_TIEMPO.xlsx</code>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    #if result.stdout:
                    #    st.text("Output:")
                    #    st.text(result.stdout)
                    
                    #else:
                    #    st.error(f"Error al generar sembrado: {result.stderr}")
                        
                except Exception as e:
                    st.error(f"Error al ejecutar el script: {e}")
    
    with col2:
        if os.path.exists("sembrado_competencia_POR_TIEMPO.xlsx"):
            st.info("📄 Archivo generado disponible para descarga")
            
            with open("sembrado_competencia_POR_TIEMPO.xlsx", "rb") as file:
                st.download_button(
                    label="⬇️ Descargar Sembrado por Tiempo",
                    data=file.read(),
                    file_name="sembrado_competencia_POR_TIEMPO.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # Sección de papeletas para jueces
            st.markdown("### 📋 Generar Papeletas para Jueces")
            st.info("Las papeletas se generan en formato Excel con rectángulos imprimibles")
            
            if st.button("📄 Generar Papeletas Excel", type="secondary", key="papeletas_excel_tiempo"):
                with st.spinner("Generando papeletas en Excel..."):
                    success, message = papeletas_excel_module.generar_papeletas_excel()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
            
            # Botón de descarga para papeletas Excel si existe
            if os.path.exists("papeletas_jueces.xlsx"):
                st.info("📄 Papeletas Excel disponibles")
                with open("papeletas_jueces.xlsx", "rb") as file:
                    st.download_button(
                        label="⬇️ Descargar Papeletas Excel",
                        data=file.read(),
                        file_name="papeletas_jueces.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

def generate_seeding_excel(df_evento, evento_nombre, tipo_sembrado):
    """
    Genera un archivo Excel con formato de sembrado actualizado con tiempos de competencia
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sembrado {tipo_sembrado.title()}"
    
    # Título principal
    ws.cell(row=1, column=1, value=evento_nombre).font = Font(bold=True, size=16)
    
    current_row = 3
    current_serie = None
    
    # Procesar por serie
    for _, row in df_evento.iterrows():
        serie_num = row['Serie']
        
        # Nueva serie
        if serie_num != current_serie:
            if current_serie is not None:
                current_row += 1  # Espacio entre series
            
            ws.cell(row=current_row, column=1, value=f"Serie {serie_num}").font = Font(bold=True, size=14)
            current_row += 1
            
            # Headers
            headers = ["Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción", "Tiempo Competencia"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                if header == "Tiempo Competencia":
                    cell.font = Font(bold=True, color="FF0000")
            current_row += 1
            current_serie = serie_num
        
        # Datos del nadador
        ws.cell(row=current_row, column=1, value=row['Carril'])
        ws.cell(row=current_row, column=2, value=row['Nombre'])
        ws.cell(row=current_row, column=3, value=row['Equipo'])
        ws.cell(row=current_row, column=4, value=row['Edad'])
        ws.cell(row=current_row, column=5, value=row['Categoría'])
        ws.cell(row=current_row, column=6, value=row['Tiempo Inscripción'])
        
        # Tiempo de competencia (editado por el usuario)
        tiempo_comp = row['Tiempo Competencia']
        comp_cell = ws.cell(row=current_row, column=7, value=tiempo_comp if tiempo_comp else "")
        comp_cell.font = Font(color="0000FF")
        
        current_row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['B'].width = 40  # Nombre
    ws.column_dimensions['C'].width = 25  # Equipo
    ws.column_dimensions['F'].width = 18  # Tiempo Inscripción
    ws.column_dimensions['G'].width = 18  # Tiempo Competencia
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

def generate_seeding_excel_from_manual(seeding_data, event_name, gender):
    """
    Genera un archivo Excel desde datos de sembrado manual
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sembrado Manual"
    
    # Título
    ws.cell(row=1, column=1, value=f"{event_name} - {gender}").font = Font(bold=True, size=16)
    
    current_row = 3
    for serie in seeding_data['series']:
        # Título de serie
        ws.cell(row=current_row, column=1, value=f"Serie {serie['serie']}").font = Font(bold=True, size=14)
        current_row += 1
        
        # Headers
        headers = ["Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción", "Tiempo Competencia"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            if header == "Tiempo Competencia":
                cell.font = Font(bold=True, color="FF0000")
        current_row += 1
        
        # Datos de carriles
        for lane_idx, swimmer in enumerate(serie['carriles']):
            ws.cell(row=current_row, column=1, value=lane_idx + 1)
            if swimmer:
                ws.cell(row=current_row, column=2, value=swimmer['nombre'])
                ws.cell(row=current_row, column=3, value=swimmer['equipo'])
                ws.cell(row=current_row, column=4, value=swimmer['edad'])
                ws.cell(row=current_row, column=5, value=swimmer['categoria'])
                ws.cell(row=current_row, column=6, value=swimmer['tiempo'])
                # Tiempo competencia (tomar valor si existe)
                comp_time = swimmer.get('tiempo_competencia', '')
                comp_cell = ws.cell(row=current_row, column=7, value=comp_time)
                comp_cell.font = Font(color="0000FF")
            current_row += 1
        
        current_row += 1  # Espacio entre series
    
    # Ajustar columnas
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

def sembrado_competencia_interface():
    st.markdown("## 📊 Sembrado de Competencia")
    
    st.markdown("""
    <div class="info-message">
        Genera los listados de participantes organizados por series y carriles para la competencia.
        Elige el método de sembrado que mejor se adapte a tu competencia.
    </div>
    """, unsafe_allow_html=True)
    
    # Verificar archivo de inscripciones
    if not os.path.exists("planilla_inscripcion.xlsx"):
        st.markdown("""
        <div class="warning-message">
            ⚠️ No se encontró el archivo <strong>planilla_inscripcion.xlsx</strong>. 
            Por favor, ve a la sección "Inscripción de Nadadores" para registrar participantes.
        </div>
        """, unsafe_allow_html=True)
        return
    
    # Verificar si hay sembrados cacheados y mostrar advertencia si puede haber cambios
    cached_seedings = []
    if 'seeding_preview_cat' in st.session_state:
        cached_seedings.append("Por Categorías")
    if 'seeding_preview_time' in st.session_state:
        cached_seedings.append("Por Tiempo")
    if any(key.startswith('manual_seeding_') for key in st.session_state.keys()):
        cached_seedings.append("Manual")
    
    if cached_seedings:
        st.markdown(f"""
        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px; padding: 12px; margin-bottom: 20px;">
            <strong>💡 Recordatorio:</strong> Tienes sembrados cargados ({', '.join(cached_seedings)}). 
            Si agregaste nuevas inscripciones, usa el botón <strong>🔄</strong> para actualizar.
        </div>
        """, unsafe_allow_html=True)
    
    # Pestañas para diferentes métodos de sembrado
    tab1, tab2, tab3 = st.tabs(["📊 Por Categorías", "⏱️ Por Tiempo", "✍️ Manual"])
    
    with tab1:
        st.markdown("### 📊 Sembrado por Categorías")
        st.markdown("""
        **¿Cuándo usar este método?**
        - Competencias federadas o oficiales
        - Eventos con múltiples categorías de edad
        - Cuando se busca competencia equitativa por grupos etarios
        
        **Cómo funciona:**
        - Agrupa nadadores por categoría de edad
        - Ordena por tiempo dentro de cada categoría
        - Coloca los mejores tiempos en las series finales
        """)
        
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("🚀 Generar Sembrado por Categorías", type="primary"):
                with st.spinner("Generando sembrado por categorías..."):
                    try:
                        script1.main_full()
                        st.markdown("""
                            <div class="success-message">
                                ✅ <strong>Sembrado generado exitosamente!</strong><br>
                                Archivo creado: <code>sembrado_competencia.xlsx</code>
                            </div>
                            """, unsafe_allow_html=True)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al generar sembrado: {e}")
        
        with col2:
            col_view, col_refresh = st.columns([2, 1])
            with col_view:
                if st.button("👁️ Visualizar Sembrado", help="Ver preview del sembrado antes de descargar"):
                    with st.spinner("Cargando visualización..."):
                        try:
                            seeding_data, message = script1.get_seeding_data()
                            if seeding_data:
                                st.session_state['seeding_preview_cat'] = seeding_data
                                st.success("✅ Visualización cargada")
                            else:
                                st.error(message)
                        except Exception as e:
                            st.error(f"Error al cargar visualización: {e}")
            
            with col_refresh:
                if st.button("🔄", help="Actualizar con nuevas inscripciones"):
                    # Limpiar cache y recargar
                    if 'seeding_preview_cat' in st.session_state:
                        del st.session_state['seeding_preview_cat']
                    with st.spinner("Actualizando sembrado..."):
                        try:
                            seeding_data, message = script1.get_seeding_data()
                            if seeding_data:
                                st.session_state['seeding_preview_cat'] = seeding_data
                                st.success("✅ Sembrado actualizado")
                            else:
                                st.error(message)
                        except Exception as e:
                            st.error(f"Error al actualizar: {e}")
        
        with col3:
            if os.path.exists("sembrado_competencia.xlsx"):
                st.info("📄 Archivo generado disponible para descarga")
                with open("sembrado_competencia.xlsx", "rb") as file:
                    st.download_button(
                        label="⬇️ Descargar Sembrado por Categorías",
                        data=file.read(),
                        file_name="sembrado_competencia.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # Mostrar visualización del sembrado si está disponible
        if 'seeding_preview_cat' in st.session_state:
            st.markdown("---")
            st.markdown("### 👁️ Vista Previa Editable - Sembrado por Categorías")
            
            seeding_data = st.session_state['seeding_preview_cat']
            
            # Selector de evento para visualizar
            eventos_disponibles = list(seeding_data.keys())
            if eventos_disponibles:
                evento_seleccionado = st.selectbox(
                    "Selecciona un evento para editar:",
                    eventos_disponibles,
                    key="evento_cat_preview"
                )
                
                if evento_seleccionado:
                    st.markdown(f"**{evento_seleccionado}**")
                    
                    # Crear una sola tabla consolidada para todas las series del evento
                    all_carriles_data = []
                    series = seeding_data[evento_seleccionado]['series']
                    
                    for serie in series:
                        for i, nadador in enumerate(serie['carriles'], 1):
                            if nadador:
                                all_carriles_data.append({
                                    "Serie": serie['serie'],
                                    "Carril": i,
                                    "Nombre": nadador['nombre'],
                                    "Equipo": nadador['equipo'],
                                    "Edad": nadador['edad'],
                                    "Categoría": nadador['categoria'],
                                    "Tiempo Inscripción": str(nadador['tiempo_inscripcion']),
                                    "Tiempo Competencia": ""
                                })
                    
                    if all_carriles_data:
                        df_evento = pd.DataFrame(all_carriles_data)
                        
                        # Tabla editable con solo la columna Tiempo Competencia editable
                        edited_df = st.data_editor(
                            df_evento,
                            disabled=["Serie", "Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción"],
                            use_container_width=True,
                            hide_index=True,
                            key=f"editor_cat_{evento_seleccionado}"
                        )
                        
                        # Botones de acción
                        col_save, col_download, col_process, col_info = st.columns([1, 1, 1, 1])
                        
                        with col_save:
                            if st.button("💾 Guardar Cambios", type="primary", help="Guardar los tiempos editados"):
                                # Actualizar session_state con los cambios
                                updated_key = f"updated_seeding_cat_{evento_seleccionado}"
                                st.session_state[updated_key] = edited_df
                                st.success("✅ Cambios guardados en memoria")
                        
                        with col_download:
                            if st.button("⬇️ Descargar Excel", help="Descargar archivo Excel con los tiempos actualizados"):
                                # Generar archivo Excel con los datos editados
                                excel_buffer = generate_seeding_excel(edited_df, evento_seleccionado, "categoria")
                                st.download_button(
                                    label="📥 Descargar Sembrado Actualizado",
                                    data=excel_buffer,
                                    file_name=f"sembrado_{evento_seleccionado.replace(' ', '_').replace('-', '_')}_editado.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        
                        with col_process:
                            if st.button("🏆 Procesar a Resultados", help="Convertir a formato de resultados para procesamiento"):
                                # Verificar si hay tiempos de competencia
                                tiempos_comp = [t for t in edited_df["Tiempo Competencia"] if t and t.strip()]
                                if len(tiempos_comp) > 0:
                                    # Guardar archivo temporalmente y procesarlo
                                    temp_file = f"temp_seeding_cat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    excel_data = generate_seeding_excel(edited_df, evento_seleccionado, "categoria")
                                    
                                    with open(temp_file, "wb") as f:
                                        f.write(excel_data)
                                    
                                    try:
                                        # Procesar con el script de resultados
                                        import importlib.util
                                        spec = importlib.util.spec_from_file_location("processor", "5-procesar_sembrado_tiempos.py")
                                        processor = importlib.util.module_from_spec(spec)
                                        spec.loader.exec_module(processor)
                                        
                                        success, message = processor.process_seeding_with_times(temp_file)
                                        
                                        if success:
                                            st.success(f"✅ {message}")
                                            st.info("📄 Archivo de resultados disponible en gestión de archivos")
                                        else:
                                            st.error(f"❌ {message}")
                                    
                                    except Exception as e:
                                        st.error(f"❌ Error al procesar: {e}")
                                    
                                    finally:
                                        # Limpiar archivo temporal
                                        if os.path.exists(temp_file):
                                            os.remove(temp_file)
                                else:
                                    st.warning("⚠️ Debes agregar al menos un tiempo de competencia")
                        
                        with col_info:
                            tiempos_completados = len([t for t in edited_df["Tiempo Competencia"] if t and t.strip()])
                            total_nadadores = len([n for n in edited_df["Nombre"] if n != "---"])
                            st.info(f"⏱️ Tiempos: {tiempos_completados}/{total_nadadores}")
    
    with tab2:
        st.markdown("### ⏱️ Sembrado por Tiempo")
        st.markdown("""
        **¿Cuándo usar este método?**
        - Competencias de clasificación o qualifiers
        - Eventos abiertos sin restricción de edad
        - Búsqueda de récords o marcas específicas
        
        **Cómo funciona:**
        - Ignora las categorías de edad
        - Ordena todos los nadadores por tiempo de inscripción
        - Series más rápidas al final del evento
        """)
        
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("🚀 Generar Sembrado por Tiempo", type="primary", key="gen_tiempo"):
                with st.spinner("Generando sembrado por tiempo..."):
                    try:
                        script2.main()
                        st.markdown("""
                            <div class="success-message">
                                ✅ <strong>Sembrado generado exitosamente!</strong><br>
                                Archivo creado: <code>sembrado_competencia_POR_TIEMPO.xlsx</code>
                            </div>
                            """, unsafe_allow_html=True)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al generar sembrado: {e}")
        
        with col2:
            col_view, col_refresh = st.columns([2, 1])
            with col_view:
                if st.button("👁️ Visualizar Sembrado", help="Ver preview del sembrado antes de descargar", key="view_tiempo"):
                    with st.spinner("Cargando visualización..."):
                        try:
                            seeding_data, message = script2.get_seeding_data()
                            if seeding_data:
                                st.session_state['seeding_preview_time'] = seeding_data
                                st.success("✅ Visualización cargada")
                            else:
                                st.error(message)
                        except Exception as e:
                            st.error(f"Error al cargar visualización: {e}")
            
            with col_refresh:
                if st.button("🔄", key="refresh_time", help="Actualizar con nuevas inscripciones"):
                    # Limpiar cache y recargar
                    if 'seeding_preview_time' in st.session_state:
                        del st.session_state['seeding_preview_time']
                    with st.spinner("Actualizando sembrado..."):
                        try:
                            seeding_data, message = script2.get_seeding_data()
                            if seeding_data:
                                st.session_state['seeding_preview_time'] = seeding_data
                                st.success("✅ Sembrado actualizado")
                            else:
                                st.error(message)
                        except Exception as e:
                            st.error(f"Error al actualizar: {e}")
        
        with col3:
            if os.path.exists("sembrado_competencia_POR_TIEMPO.xlsx"):
                st.info("📄 Archivo generado disponible para descarga")
                with open("sembrado_competencia_POR_TIEMPO.xlsx", "rb") as file:
                    st.download_button(
                        label="⬇️ Descargar Sembrado por Tiempo",
                        data=file.read(),
                        file_name="sembrado_competencia_POR_TIEMPO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # Mostrar visualización del sembrado si está disponible
        if 'seeding_preview_time' in st.session_state:
            st.markdown("---")
            st.markdown("### 👁️ Vista Previa Editable - Sembrado por Tiempo")
            
            seeding_data = st.session_state['seeding_preview_time']
            
            # Selector de evento para visualizar
            eventos_disponibles = list(seeding_data.keys())
            if eventos_disponibles:
                evento_seleccionado = st.selectbox(
                    "Selecciona un evento para editar:",
                    eventos_disponibles,
                    key="evento_time_preview"
                )
                
                if evento_seleccionado:
                    st.markdown(f"**{evento_seleccionado}**")
                    
                    # Crear una sola tabla consolidada para todas las series del evento
                    all_carriles_data = []
                    series = seeding_data[evento_seleccionado]['series']
                    
                    for serie in series:
                        for i, nadador in enumerate(serie['carriles'], 1):
                            if nadador:
                                all_carriles_data.append({
                                    "Serie": serie['serie'],
                                    "Carril": i,
                                    "Nombre": nadador['nombre'],
                                    "Equipo": nadador['equipo'],
                                    "Edad": nadador['edad'],
                                    "Categoría": nadador['categoria'],
                                    "Tiempo Inscripción": str(nadador['tiempo_inscripcion']),
                                    "Tiempo Competencia": ""
                                })
                    
                    if all_carriles_data:
                        df_evento = pd.DataFrame(all_carriles_data)
                        
                        # Tabla editable con solo la columna Tiempo Competencia editable
                        edited_df = st.data_editor(
                            df_evento,
                            disabled=["Serie", "Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción"],
                            use_container_width=True,
                            hide_index=True,
                            key=f"editor_time_{evento_seleccionado}"
                        )
                        
                        # Botones de acción
                        col_save, col_download, col_process, col_info = st.columns([1, 1, 1, 1])
                        
                        with col_save:
                            if st.button("💾 Guardar Cambios", type="primary", help="Guardar los tiempos editados", key="save_time"):
                                # Actualizar session_state con los cambios
                                updated_key = f"updated_seeding_time_{evento_seleccionado}"
                                st.session_state[updated_key] = edited_df
                                st.success("✅ Cambios guardados en memoria")
                        
                        with col_download:
                            if st.button("⬇️ Descargar Excel", help="Descargar archivo Excel con los tiempos actualizados", key="download_time"):
                                # Generar archivo Excel con los datos editados
                                excel_buffer = generate_seeding_excel(edited_df, evento_seleccionado, "tiempo")
                                st.download_button(
                                    label="📥 Descargar Sembrado Actualizado",
                                    data=excel_buffer,
                                    file_name=f"sembrado_{evento_seleccionado.replace(' ', '_').replace('-', '_')}_editado.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_time"
                                )
                        
                        with col_process:
                            if st.button("🏆 Procesar a Resultados", help="Convertir a formato de resultados para procesamiento", key="process_time"):
                                # Verificar si hay tiempos de competencia
                                tiempos_comp = [t for t in edited_df["Tiempo Competencia"] if t and t.strip()]
                                if len(tiempos_comp) > 0:
                                    # Guardar archivo temporalmente y procesarlo
                                    temp_file = f"temp_seeding_time_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    excel_data = generate_seeding_excel(edited_df, evento_seleccionado, "tiempo")
                                    
                                    with open(temp_file, "wb") as f:
                                        f.write(excel_data)
                                    
                                    try:
                                        # Procesar con el script de resultados
                                        import importlib.util
                                        spec = importlib.util.spec_from_file_location("processor", "5-procesar_sembrado_tiempos.py")
                                        processor = importlib.util.module_from_spec(spec)
                                        spec.loader.exec_module(processor)
                                        
                                        success, message = processor.process_seeding_with_times(temp_file)
                                        
                                        if success:
                                            st.success(f"✅ {message}")
                                            st.info("📄 Archivo de resultados disponible en gestión de archivos")
                                        else:
                                            st.error(f"❌ {message}")
                                    
                                    except Exception as e:
                                        st.error(f"❌ Error al procesar: {e}")
                                    
                                    finally:
                                        # Limpiar archivo temporal
                                        if os.path.exists(temp_file):
                                            os.remove(temp_file)
                                else:
                                    st.warning("⚠️ Debes agregar al menos un tiempo de competencia")
                        
                        with col_info:
                            tiempos_completados = len([t for t in edited_df["Tiempo Competencia"] if t and t.strip()])
                            total_nadadores = len([n for n in edited_df["Nombre"] if n != "---"])
                            st.info(f"⏱️ Tiempos: {tiempos_completados}/{total_nadadores}")
        
    
    with tab3:
        st.markdown("### ✍️ Sembrado Manual")
        st.markdown("""
        **Crea sembrados personalizados con total control:**
        - 🎯 Arrastra nadadores entre carriles y series
        - 🔄 Combina diferentes categorías  
        - 📊 Vista en tiempo real del sembrado
        - 💾 Guarda tu organización personalizada
        """)
        
        # Verificar si hay inscripciones
        if not os.path.exists("planilla_inscripcion.xlsx"):
            st.warning("⚠️ Necesitas tener nadadores inscritos para crear un sembrado manual.")
            st.info("👉 Ve a la sección **Inscripción de Nadadores** para registrar participantes primero.")
            return
        
        try:
            # Cargar datos de inscripción
            df = pd.read_excel("planilla_inscripcion.xlsx")
            info_cols = ['NOMBRE Y AP', 'EQUIPO', 'EDAD', 'CAT.', 'SEXO']
            event_cols = [col for col in df.columns if col not in info_cols and 'Nø' not in col and 'FECHA DE NA' not in col]
            
            if len(event_cols) == 0:
                st.error("❌ No se encontraron eventos en la planilla de inscripción")
                return
                
        except Exception as e:
            st.error(f"❌ Error al leer inscripciones: {e}")
            return
        
        # Selector de evento
        st.markdown("#### 📋 Seleccionar Evento")
        col_event, col_gender = st.columns([2, 1])
        
        with col_event:
            selected_event = st.selectbox(
                "Selecciona el evento para crear sembrado manual:",
                event_cols,
                key="manual_event_select"
            )
        
        with col_gender:
            gender_filter = st.selectbox(
                "Género:",
                ["Todos", "Masculino", "Femenino"],
                key="manual_gender_filter"
            )
        
        if selected_event:
            # Filtrar nadadores para el evento seleccionado
            swimmers_for_event = []
            for index, row in df.iterrows():
                if inscrito_en_prueba(row[selected_event]) and pd.notna(row['NOMBRE Y AP']):
                    swimmer_gender = "Masculino" if row['SEXO'].upper() == 'M' else "Femenino"
                    if gender_filter == "Todos" or gender_filter == swimmer_gender:
                        swimmers_for_event.append({
                            'id': index,
                            'nombre': row['NOMBRE Y AP'],
                            'equipo': row['EQUIPO'],
                            'edad': row['EDAD'],
                            'categoria': row['CAT.'],
                            'sexo': swimmer_gender,
                            'tiempo': str(row[selected_event]),
                            'tiempo_en_segundos': script1.parse_time(row[selected_event]),
                        })
            
            if len(swimmers_for_event) == 0:
                st.warning(f"⚠️ No hay nadadores inscritos en {selected_event} con el filtro seleccionado")
                return
            
            st.success(f"✅ {len(swimmers_for_event)} nadadores encontrados en **{selected_event}**")
            
            # Botón para actualizar sembrado con nuevas inscripciones
            col_refresh, col_info = st.columns([1, 3])
            with col_refresh:
                if st.button("🔄 Actualizar Sembrado", help="Cargar nuevas inscripciones"):
                    seeding_key = f"manual_seeding_{selected_event}_{gender_filter}"
                    if seeding_key in st.session_state:
                        del st.session_state[seeding_key]
                    st.rerun()
            
            with col_info:
                st.info("💡 Usa 'Actualizar Sembrado' si agregaste nuevas inscripciones")

            # Inicializar sembrado en session state
            seeding_key = f"manual_seeding_{selected_event}_{gender_filter}"
            
            def create_initial_seeding(swimmers_list):
                """Crear sembrado inicial automático"""
                sorted_swimmers = sorted(
                    swimmers_list,
                    key=lambda x: x.get('tiempo_en_segundos', script1.parse_time(x['tiempo'])),
                )
                
                # Distribución en series de 8 carriles
                series = []
                swimmers_per_series = 8
                num_series = (len(sorted_swimmers) + swimmers_per_series - 1) // swimmers_per_series
                
                for serie_num in range(num_series):
                    serie_swimmers = sorted_swimmers[serie_num * swimmers_per_series:(serie_num + 1) * swimmers_per_series]
                    lane_assignment = [4, 5, 3, 6, 2, 7, 1, 8]  # Orden standard de carriles
                    
                    serie = {"serie": serie_num + 1, "carriles": [None] * 8}
                    for i, swimmer in enumerate(serie_swimmers):
                        if i < len(lane_assignment):
                            lane_idx = lane_assignment[i] - 1
                            serie["carriles"][lane_idx] = swimmer
                    
                    series.append(serie)
                
                return {
                    'evento': selected_event,
                    'genero': gender_filter, 
                    'series': series,
                    'nadadores_disponibles': [],
                    'total_nadadores': len(swimmers_list)  # Para detectar cambios
                }
            
            # Verificar si necesita actualización automática
            if seeding_key not in st.session_state:
                st.session_state[seeding_key] = create_initial_seeding(swimmers_for_event)
            else:
                # Verificar si el número de nadadores cambió
                current_total = st.session_state[seeding_key].get('total_nadadores', 0)
                if current_total != len(swimmers_for_event):
                    st.warning(f"⚠️ Se detectaron {len(swimmers_for_event) - current_total} nuevas inscripciones. Usa 'Actualizar Sembrado' para cargarlas.")
            
            seeding_data = st.session_state[seeding_key]
            
            # Interfaz de edición manual
            st.markdown("#### 🎯 Editor de Sembrado Manual")
            
            # Mostrar nadadores disponibles (no asignados)
            if seeding_data['nadadores_disponibles']:
                st.markdown("##### 👥 Nadadores Disponibles")
                cols_available = st.columns(min(len(seeding_data['nadadores_disponibles']), 4))
                category_colors = {
                    'MENORES': '#FFB6C1',  'JUVENIL A': '#87CEEB',  'JUVENIL B': '#98FB98',
                    'JUNIOR': '#DDA0DD',    'SENIOR': '#F0E68C',    'MASTER': '#FFA07A'
                }
                for i, swimmer in enumerate(seeding_data['nadadores_disponibles']):
                    with cols_available[i % 4]:
                        bg_color = category_colors.get(swimmer['categoria'], '#E6E6FA')
                        st.markdown(f"""
                        <div style="background-color: {bg_color}; padding: 6px; border-radius: 4px; border: 1px solid #ccc; margin: 2px 0;">
                            <strong>{swimmer['nombre']}</strong><br>
                            <small>{swimmer['equipo']} - {swimmer['categoria']}</small>
                        </div>
                        """, unsafe_allow_html=True)
            
            # Leyenda de colores
            st.markdown("##### 🎨 Leyenda de Categorías")
            category_colors = {
                'MENORES': '#FFB6C1',  'JUVENIL A': '#87CEEB',  'JUVENIL B': '#98FB98',
                'JUNIOR': '#DDA0DD',    'SENIOR': '#F0E68C',    'MASTER': '#FFA07A'
            }
            
            legend_cols = st.columns(len(category_colors))
            for i, (cat, color) in enumerate(category_colors.items()):
                with legend_cols[i]:
                    st.markdown(f"""
                    <div style="background-color: {color}; padding: 4px; border-radius: 3px; text-align: center; border: 1px solid #ccc;">
                        <small><strong>{cat}</strong></small>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Editor por series
            st.markdown("##### 🏊 Series y Carriles")
            
            for serie_idx, serie in enumerate(seeding_data['series']):
                st.markdown(f"**Serie {serie['serie']}**")
                
                # Crear columnas para los 8 carriles
                lane_cols = st.columns(8)
                
                for lane_idx in range(8):
                    with lane_cols[lane_idx]:
                        st.markdown(f"**Carril {lane_idx + 1}**")
                        
                        current_swimmer = serie['carriles'][lane_idx]
                        
                        if current_swimmer:
                            # Mostrar nadador actual con color por categoría
                            category_colors = {
                                'MENORES': '#FFB6C1',  # Rosa claro
                                'JUVENIL A': '#87CEEB',  # Azul cielo
                                'JUVENIL B': '#98FB98',  # Verde claro
                                'JUNIOR': '#DDA0DD',    # Violeta claro
                                'SENIOR': '#F0E68C',    # Amarillo claro
                                'MASTER': '#FFA07A'     # Salmón
                            }
                            bg_color = category_colors.get(current_swimmer['categoria'], '#E6E6FA')
                            
                            st.markdown(f"""
                            <div style="background-color: {bg_color}; padding: 8px; border-radius: 5px; border: 1px solid #ccc; margin: 2px 0;">
                                <strong>{current_swimmer['nombre']}</strong><br>
                                <small>{current_swimmer['equipo']} | {current_swimmer['categoria']}</small>
                            </div>
                            """, unsafe_allow_html=True)

                            # Mostrar tiempo de sembrado (no editable)
                            st.text_input(
                                "T. Sembrado:",
                                value=current_swimmer['tiempo'],
                                key=f"seeding_time_{serie_idx}_{lane_idx}",
                                disabled=True,
                                help="Tiempo original de sembrado (no modificable)"
                            )

                            # Campo editable para tiempo de competencia
                            competition_time_key = f"comp_time_{serie_idx}_{lane_idx}"
                            current_comp_time = current_swimmer.get('tiempo_competencia', '')

                            new_comp_time = st.text_input(
                                "T. Competencia:",
                                value=current_comp_time,
                                key=competition_time_key,
                                placeholder="MM:SS.dd",
                                help="Ingresa el tiempo de competencia en formato MM:SS.dd (ej: 02:15.45)"
                            )

                            # Actualizar tiempo de competencia si cambió
                            if new_comp_time != current_comp_time:
                                seeding_data['series'][serie_idx]['carriles'][lane_idx]['tiempo_competencia'] = new_comp_time
                                st.session_state[seeding_key] = seeding_data

                            # Botón para remover nadador
                            if st.button("❌", key=f"remove_{serie_idx}_{lane_idx}", help="Remover nadador"):
                                # Mover a disponibles
                                seeding_data['nadadores_disponibles'].append(current_swimmer)
                                seeding_data['series'][serie_idx]['carriles'][lane_idx] = None
                                st.rerun()
                        else:
                            # Carril vacío - mostrar opciones para asignar
                            st.info("🔘 Carril vacío")
                            
                            # Crear lista de nadadores para asignar
                            available_swimmers = seeding_data['nadadores_disponibles'].copy()
                            
                            # Agregar nadadores de otros carriles para intercambio
                            for s_idx, s in enumerate(seeding_data['series']):
                                for l_idx, swimmer in enumerate(s['carriles']):
                                    if swimmer and not (s_idx == serie_idx and l_idx == lane_idx):
                                        available_swimmers.append({**swimmer, '_from_serie': s_idx, '_from_lane': l_idx})
                            
                            if available_swimmers:
                                swimmer_options = ["Seleccionar nadador..."] + [
                                    f"{swimmer['nombre']} ({swimmer['equipo']} - {swimmer['categoria']})" 
                                    for swimmer in available_swimmers
                                ]
                                
                                selected_swimmer_idx = st.selectbox(
                                    "Asignar:",
                                    range(len(swimmer_options)),
                                    format_func=lambda x: swimmer_options[x],
                                    key=f"assign_{serie_idx}_{lane_idx}"
                                )
                                
                                if selected_swimmer_idx > 0:
                                    selected_swimmer = available_swimmers[selected_swimmer_idx - 1]
                                    
                                    if st.button("✅ Asignar", key=f"confirm_{serie_idx}_{lane_idx}"):
                                        # Asignar nadador al carril
                                        if '_from_serie' in selected_swimmer:
                                            # Mover de otro carril (intercambio)
                                            from_serie = selected_swimmer['_from_serie']
                                            from_lane = selected_swimmer['_from_lane']
                                            clean_swimmer = {k: v for k, v in selected_swimmer.items() if not k.startswith('_')}
                                            
                                            seeding_data['series'][serie_idx]['carriles'][lane_idx] = clean_swimmer
                                            seeding_data['series'][from_serie]['carriles'][from_lane] = None
                                        else:
                                            # Mover de disponibles
                                            seeding_data['series'][serie_idx]['carriles'][lane_idx] = selected_swimmer
                                            seeding_data['nadadores_disponibles'].remove(selected_swimmer)
                                        
                                        st.rerun()
                
                st.markdown("---")
            
            # Controles adicionales
            st.markdown("#### ⚙️ Controles Avanzados")
            col_add_series, col_category_info = st.columns([1, 2])
            
            with col_add_series:
                col_add, col_remove = st.columns([1, 1])
                with col_add:
                    if st.button("➕ Nueva Serie"):
                        new_serie_num = len(seeding_data['series']) + 1
                        new_serie = {"serie": new_serie_num, "carriles": [None] * 8}
                        seeding_data['series'].append(new_serie)
                        st.rerun()
                
                with col_remove:
                    if len(seeding_data['series']) > 1:
                        if st.button("🗑️ Eliminar Serie", help="Eliminar última serie vacía"):
                            # Verificar si la última serie está vacía
                            last_serie = seeding_data['series'][-1]
                            if all(swimmer is None for swimmer in last_serie['carriles']):
                                seeding_data['series'].pop()
                                st.rerun()
                            else:
                                st.error("❌ Solo se pueden eliminar series vacías")
            
            with col_category_info:
                # Mostrar distribución de categorías
                categories_in_seeding = {}
                total_swimmers = 0
                for serie in seeding_data['series']:
                    for swimmer in serie['carriles']:
                        if swimmer:
                            cat = swimmer['categoria']
                            categories_in_seeding[cat] = categories_in_seeding.get(cat, 0) + 1
                            total_swimmers += 1
                
                if categories_in_seeding:
                    st.info(f"📊 **Nadadores por categoría:** " + 
                           ", ".join([f"{cat}: {count}" for cat, count in sorted(categories_in_seeding.items())]) +
                           f" | **Total: {total_swimmers}**")
            
            # Botones de acción
            st.markdown("#### 💾 Acciones")
            col_save, col_download, col_reset = st.columns([1, 1, 1])
            
            with col_save:
                if st.button("💾 Guardar Sembrado", type="primary"):
                    # Generar archivo Excel con sembrado manual
                    manual_filename = f"sembrado_manual_{selected_event}_{gender_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sembrado Manual"
                    
                    # Título
                    ws.cell(row=1, column=1, value=f"{selected_event} - {gender_filter}").font = Font(bold=True, size=16)
                    
                    current_row = 3
                    for serie in seeding_data['series']:
                        # Título de serie
                        ws.cell(row=current_row, column=1, value=f"Serie {serie['serie']}").font = Font(bold=True, size=14)
                        current_row += 1
                        
                        # Headers
                        headers = ["Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción", "Tiempo Competencia"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col, value=header)
                            cell.font = Font(bold=True)
                            if header == "Tiempo Competencia":
                                cell.font = Font(bold=True, color="FF0000")
                        current_row += 1
                        
                        # Datos de carriles
                        for lane_idx, swimmer in enumerate(serie['carriles']):
                            ws.cell(row=current_row, column=1, value=lane_idx + 1)
                            if swimmer:
                                ws.cell(row=current_row, column=2, value=swimmer['nombre'])
                                ws.cell(row=current_row, column=3, value=swimmer['equipo'])
                                ws.cell(row=current_row, column=4, value=swimmer['edad'])
                                ws.cell(row=current_row, column=5, value=swimmer['categoria'])
                                ws.cell(row=current_row, column=6, value=swimmer['tiempo'])
                                comp_cell = ws.cell(row=current_row, column=7, value="")
                                comp_cell.font = Font(color="0000FF")
                            current_row += 1
                        
                        current_row += 1  # Espacio entre series
                    
                    # Ajustar columnas
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['F'].width = 18
                    ws.column_dimensions['G'].width = 18
                    
                    wb.save(manual_filename)
                    st.success(f"✅ Sembrado manual guardado: {manual_filename}")
            
            with col_download:
                if st.button("⬇️ Descargar Excel"):
                    # Crear archivo temporal para descarga
                    manual_buffer = generate_seeding_excel_from_manual(seeding_data, selected_event, gender_filter)
                    st.download_button(
                        label="📥 Descargar Sembrado Manual",
                        data=manual_buffer,
                        file_name=f"sembrado_manual_{selected_event}_{gender_filter}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            with col_reset:
                if st.button("🔄 Resetear", help="Volver al sembrado automático inicial"):
                    if seeding_key in st.session_state:
                        del st.session_state[seeding_key]
                    st.rerun()
    
    # Sección de limpieza de sembrados
    st.markdown("---")
    st.markdown("### 🧹 Limpiar Sembrados")
    
    col_clean_sem1, col_clean_sem2, col_clean_sem3 = st.columns([2, 1, 1])
    
    with col_clean_sem1:
        st.info("🗑️ Eliminar archivos de sembrado para generar nuevos")
    
    with col_clean_sem2:
        if st.button("📊 Limpiar Por Categoría", help="Eliminar sembrado por categoría"):
            if os.path.exists("sembrado_competencia.xlsx"):
                try:
                    os.remove("sembrado_competencia.xlsx")
                    # Limpiar session state relacionado
                    if 'seeding_preview_cat' in st.session_state:
                        del st.session_state['seeding_preview_cat']
                    st.success("✅ Sembrado por categoría eliminado")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")
            else:
                st.info("ℹ️ No hay sembrado por categoría")
    
    with col_clean_sem3:
        if st.button("⏱️ Limpiar Por Tiempo", help="Eliminar sembrado por tiempo"):
            if os.path.exists("sembrado_competencia_POR_TIEMPO.xlsx"):
                try:
                    os.remove("sembrado_competencia_POR_TIEMPO.xlsx")
                    # Limpiar session state relacionado
                    if 'seeding_preview_time' in st.session_state:
                        del st.session_state['seeding_preview_time']
                    st.success("✅ Sembrado por tiempo eliminado")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")
            else:
                st.info("ℹ️ No hay sembrado por tiempo")

def procesar_resultados():
    st.markdown("## 🏆 Procesar Resultados")

    st.markdown("""
    <div class="info-message">
        Procesa los resultados finales tomando los tiempos de competencia del sembrado manual
        y genera reportes de premiación con sistema de puntos y clasificaciones por categoría, género y equipos.
    </div>
    """, unsafe_allow_html=True)

    # Verificar si hay datos de sembrado manual con tiempos de competencia
    tiempos_competencia_disponibles = False
    total_tiempos = 0

    for key in st.session_state.keys():
        if key.startswith('manual_seeding_'):
            seeding_data = st.session_state[key]
            for serie in seeding_data.get('series', []):
                for swimmer in serie['carriles']:
                    if swimmer and swimmer.get('tiempo_competencia'):
                        tiempos_competencia_disponibles = True
                        total_tiempos += 1

    if not tiempos_competencia_disponibles:
        st.markdown("""
        <div class="warning-message">
            ⚠️ No se encontraron tiempos de competencia en el sembrado manual.
            Por favor, ingresa los tiempos de competencia en la sección <strong>"Sembrado Manual"</strong>.
        </div>
        """, unsafe_allow_html=True)
        return

    st.success(f"✅ Se encontraron {total_tiempos} tiempos de competencia listos para procesar")

    # Mostrar nuevo sistema de puntos
    st.markdown("### 🎯 Nuevo Sistema de Puntos")

    col_puntos1, col_puntos2 = st.columns(2)

    with col_puntos1:
        st.markdown("**Primeros lugares:**")
        puntos_principales = pd.DataFrame({
            'Posición': [1, 2, 3, 4],
            'Puntos': [9, 7, 6, 5]
        })
        st.dataframe(puntos_principales, use_container_width=True, hide_index=True)

    with col_puntos2:
        st.markdown("**Lugares restantes:**")
        puntos_restantes = pd.DataFrame({
            'Posición': [5, 6, 7, 8, '9+'],
            'Puntos': [4, 3, 2, 1, '1*']
        })
        st.dataframe(puntos_restantes, use_container_width=True, hide_index=True)

    st.markdown("_*A partir del 5º lugar se resta 1 punto por posición (mínimo 1 punto)_")

    col1, col2 = st.columns([1, 2])

    with col1:
        if st.button("🚀 Procesar Resultados", type="primary"):
            with st.spinner("Procesando resultados con tiempos de competencia..."):
                try:
                    success, message = script3.generar_reporte_resultados_completo()
                    if success:
                        st.markdown(f"""
                            <div class="success-message">
                                ✅ <strong>Resultados procesados exitosamente!</strong><br>
                                {message}
                            </div>
                            """, unsafe_allow_html=True)
                        st.rerun()
                    else:
                        st.error(f"Error al procesar resultados: {message}")
                except Exception as e:
                    st.error(f"Error al procesar resultados: {e}")

    with col2:
        if st.button("👁️ Vista Previa de Resultados", help="Ver vista previa de los resultados antes de procesar"):
            with st.spinner("Cargando vista previa..."):
                try:
                    resultados_brutos = script3.leer_tiempos_competencia_desde_sembrado()
                    if resultados_brutos:
                        resultados_procesados = script3.procesar_resultados_por_categoria_y_genero(resultados_brutos)
                        st.session_state['resultados_preview'] = resultados_procesados
                        st.success(f"✅ Vista previa cargada: {len(resultados_procesados)} resultados")
                    else:
                        st.error("No se encontraron resultados para procesar")
                except Exception as e:
                    st.error(f"Error al cargar vista previa: {e}")
    
    # Botón de descarga
    if os.path.exists("reporte_premiacion_final_CORREGIDO.xlsx"):
        st.info("📄 Reporte de premiación disponible para descarga")
        with open("reporte_premiacion_final_CORREGIDO.xlsx", "rb") as file:
            st.download_button(
                label="⬇️ Descargar Reporte de Premiación Excel",
                data=file.read(),
                file_name="reporte_premiacion_final_CORREGIDO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Mostrar vista previa de resultados si están disponibles
    if 'resultados_preview' in st.session_state:
        st.markdown("---")
        st.markdown("### 📊 Vista Previa de Resultados")

        resultados_data = st.session_state['resultados_preview']

        if resultados_data:
            df_resultados = pd.DataFrame(resultados_data)

            tab1, tab2 = st.tabs(["🏅 Resultados por Categoría y Género", "🏢 Resumen por Equipos"])

            with tab1:
                st.markdown("#### 🏅 Ranking por Categoría y Género (Ordenado por Tiempo)")

                # Agrupar por evento, género y categoría
                for (evento, genero, categoria), grupo in df_resultados.groupby(['Evento', 'Genero', 'Categoria']):
                    with st.expander(f"🏊 {evento} - {genero} - {categoria} ({len(grupo)} nadadores)"):
                        df_display = grupo[['Posicion', 'Nombre', 'Equipo', 'Edad', 'Tiempo_Competencia', 'Puntos']].copy()
                        df_display.columns = ['Pos.', 'Nombre', 'Equipo', 'Edad', 'Tiempo', 'Puntos']

                        # Función para destacar ganadores
                        def highlight_winners(row):
                            if row['Pos.'] == 1:
                                return ['background-color: #FFD700; color: black'] * len(row)  # Oro
                            elif row['Pos.'] == 2:
                                return ['background-color: #C0C0C0; color: black'] * len(row)  # Plata
                            elif row['Pos.'] == 3:
                                return ['background-color: #CD7F32; color: white'] * len(row)  # Bronce
                            else:
                                return [''] * len(row)

                        st.dataframe(
                            df_display.style.apply(highlight_winners, axis=1),
                            use_container_width=True,
                            hide_index=True
                        )

            with tab2:
                st.markdown("#### 🏢 Resumen de Puntos por Equipo")

                # Calcular puntos por equipo
                puntos_equipos = df_resultados.groupby('Equipo').agg({
                    'Puntos': 'sum',
                    'Nombre': 'count'
                }).reset_index()
                puntos_equipos.columns = ['Equipo', 'Puntos Totales', 'Participaciones']
                puntos_equipos = puntos_equipos.sort_values('Puntos Totales', ascending=False)
                puntos_equipos.reset_index(drop=True, inplace=True)
                puntos_equipos.index += 1  # Empezar desde 1

                # Función para destacar primeros equipos
                def highlight_teams(row):
                    if row.name == 1:
                        return ['background-color: #FFD700; color: black'] * len(row)  # Oro
                    elif row.name == 2:
                        return ['background-color: #C0C0C0; color: black'] * len(row)  # Plata
                    elif row.name == 3:
                        return ['background-color: #CD7F32; color: white'] * len(row)  # Bronce
                    else:
                        return [''] * len(row)

                st.dataframe(
                    puntos_equipos.style.apply(highlight_teams, axis=1),
                    use_container_width=True
                )
    
    # Sección de limpieza de resultados
    st.markdown("---")
    st.markdown("### 🧹 Limpiar Resultados")
    
    col_clean_res1, col_clean_res2, col_clean_res3 = st.columns([2, 1, 1])
    
    with col_clean_res1:
        st.info("🗑️ Eliminar archivos de resultados y reportes")
    
    with col_clean_res2:
        if st.button("🏆 Limpiar Resultados", help="Eliminar archivo de resultados de competencia"):
            files_to_clean = ["resultados_con_tiempos.xlsx"]
            import glob
            dynamic_results = glob.glob("resultados_desde_sembrado_*.xlsx")
            files_to_clean.extend(dynamic_results)
            
            deleted = []
            for file in files_to_clean:
                if os.path.exists(file):
                    try:
                        os.remove(file)
                        deleted.append(file)
                    except Exception as e:
                        st.error(f"❌ Error eliminando {file}: {e}")
            
            if deleted:
                st.success(f"✅ Eliminados: {', '.join(deleted)}")
                st.rerun()
            else:
                st.info("ℹ️ No hay resultados que eliminar")
    
    with col_clean_res3:
        if st.button("🏅 Limpiar Reportes", help="Eliminar reporte de premiación"):
            if os.path.exists("reporte_premiacion_final_CORREGIDO.xlsx"):
                try:
                    os.remove("reporte_premiacion_final_CORREGIDO.xlsx")
                    st.success("✅ Reporte de premiación eliminado")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")
            else:
                st.info("ℹ️ No hay reporte que eliminar")

def generar_papeletas_interface():
    """Interfaz independiente para generar papeletas PDF y Excel con vista previa"""
    st.markdown("## 📋 Generar Papeletas para Jueces")
    
    st.markdown("""
    <div class="info-message">
        <p>Genera papeletas <strong>individuales por nadador</strong> para que los jueces registren los tiempos durante la competencia.</p>
        <p><strong>Requisito:</strong> Debes tener el archivo de inscripción cargado.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Verificar archivo de entrada
    if not os.path.exists("planilla_inscripcion.xlsx"):
        st.markdown("""
        <div class="warning-message">
            ⚠️ No se encontró el archivo <strong>planilla_inscripcion.xlsx</strong>. 
            Por favor, súbelo en la sección "Gestión de Archivos" o registra nadadores primero.
        </div>
        """, unsafe_allow_html=True)
        return
    
    # Leer y mostrar vista previa de papeletas
    try:
        papeletas_data = papeletas_excel_module.leer_datos_sembrado()
        if not papeletas_data:
            st.error("No se encontraron datos del sembrado")
            return
        
        st.success(f"✅ Se generarán {len(papeletas_data)} papeletas individuales")
        
        # Vista previa de papeletas
        st.markdown("### 👁️ Vista Previa de Papeletas")
        
        # Navegador de papeletas
        col_nav1, col_nav2, col_nav3 = st.columns([1, 2, 1])
        
        with col_nav2:
            # Selector de papeleta
            papeleta_index = st.selectbox(
                f"Selecciona papeleta (1-{len(papeletas_data)}):",
                range(len(papeletas_data)),
                format_func=lambda x: f"Papeleta {x+1}: {papeletas_data[x]['nombre']} - {papeletas_data[x]['prueba'].split(' - ')[0]}",
                key="papeleta_selector"
            )
        
        with col_nav1:
            if st.button("⬅️ Anterior", disabled=(papeleta_index == 0)):
                if papeleta_index > 0:
                    st.session_state.papeleta_selector = papeleta_index - 1
                    st.rerun()
        
        with col_nav3:
            if st.button("Siguiente ➡️", disabled=(papeleta_index == len(papeletas_data)-1)):
                if papeleta_index < len(papeletas_data)-1:
                    st.session_state.papeleta_selector = papeleta_index + 1
                    st.rerun()
        
        # Mostrar papeleta seleccionada
        nadador_actual = papeletas_data[papeleta_index]
        
        # Mostrar papeleta usando componentes nativos de Streamlit
        with st.container():
            st.markdown("---")
            
            # Título de la prueba
            st.markdown(f"### 🏊‍♂️ PRUEBA: {nadador_actual['prueba']}")
            
            # Información del nadador
            st.markdown(f"**👤 Nadador:** {nadador_actual['nombre']}")
            st.markdown(f"**🏢 Equipo:** {nadador_actual['equipo']}")
            st.markdown(f"**📊 Categoría:** {nadador_actual['categoria']}")
            
            # Serie y Carril en columnas
            col_serie, col_carril = st.columns(2)
            with col_serie:
                st.metric("🏁 SERIE", nadador_actual['serie'])
            with col_carril:
                st.metric("🛤️ CARRIL", nadador_actual['carril'])
            
            # Sección de tiempo prominente
            st.markdown("---")
            st.markdown("### ⏱️ TIEMPO DE COMPETENCIA")
            
            # Crear un espacio visual para el tiempo con CSS más simple
            st.markdown("""
            <div style="
                text-align: center; 
                font-size: 24px; 
                font-weight: bold; 
                font-family: monospace;
                border: 3px solid #000000; 
                padding: 20px; 
                margin: 10px 0;
                background-color: #ffffff;
                color: #000000;
                letter-spacing: 4px;
            ">
                _____ : _____ . _____
            </div>
            """, unsafe_allow_html=True)
            
            # Información del juez
            st.markdown("---")
            st.markdown("**👨‍⚖️ Juez:** ________________________________")
            
            col_fecha, col_hora = st.columns(2)
            with col_fecha:
                st.markdown("**📅 Fecha:** ______________")
            with col_hora:
                st.markdown("**🕐 Hora:** ______________")
            
            st.markdown("---")
        
        # Información adicional de la papeleta actual
        col_info1, col_info2, col_info3, col_info4 = st.columns(4)
        with col_info1:
            st.metric("Papeleta", f"{papeleta_index + 1} de {len(papeletas_data)}")
        with col_info2:
            st.metric("Serie Asignada", f"{nadador_actual['serie']}")
        with col_info3:
            st.metric("Carril Asignado", f"{nadador_actual['carril']}")
        with col_info4:
            st.metric("Tiempo Inscripción", f"{nadador_actual['tiempo_inscripcion']}")
        
        # Botón para vista rápida de impresión
        if st.button("🖨️ Vista de Impresión", key="print_preview"):
            st.markdown("### 🖨️ Vista de Impresión - Simulación Tamaño Real")
            
            # Vista de impresión usando componentes Streamlit nativos
            with st.container():
                # Crear un borde visual con CSS simple
                st.markdown("""
                <div style="
                    border: 2px solid #000000; 
                    padding: 20px; 
                    margin: 10px; 
                    background-color: #ffffff;
                ">
                """, unsafe_allow_html=True)
                
                # Contenido de la papeleta
                st.markdown(f"## 🏊‍♂️ PRUEBA: {nadador_actual['prueba']}")
                st.markdown("---")
                
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.markdown(f"### 👤 {nadador_actual['nombre']}")
                    st.markdown(f"**🏢 {nadador_actual['equipo']} - 📊 {nadador_actual['categoria']}**")
                
                st.markdown("---")
                
                # Serie y Carril grandes
                col_s, col_c = st.columns(2)
                with col_s:
                    st.markdown("#### 🏁 SERIE:")
                    st.markdown(f"# {nadador_actual['serie']}")
                with col_c:
                    st.markdown("#### 🛤️ CARRIL:")
                    st.markdown(f"# {nadador_actual['carril']}")
                
                st.markdown("---")
                st.markdown("## ⏱️ TIEMPO DE COMPETENCIA:")
                
                # Tiempo en formato grande
                st.markdown("""
                <div style="
                    text-align: center; 
                    font-size: 36px; 
                    font-weight: bold; 
                    font-family: monospace;
                    border: 4px solid #000000; 
                    padding: 30px; 
                    margin: 20px 0;
                    background-color: #f9f9f9;
                    letter-spacing: 8px;
                ">
                    _____ : _____ . _____
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("---")
                st.markdown("**👨‍⚖️ Juez:** ________________________________")
                
                col_f, col_h = st.columns(2)
                with col_f:
                    st.markdown("**📅 Fecha:** ______________")
                with col_h:
                    st.markdown("**🕐 Hora:** ______________")
                
                # Cerrar el div del borde
                st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("*Esta vista simula el tamaño real de impresión*")
        
        # Lista compacta de todas las papeletas
        with st.expander(f"📊 Ver lista completa ({len(papeletas_data)} papeletas)"):
            df_preview = pd.DataFrame(papeletas_data)
            st.dataframe(
                df_preview[['nombre', 'prueba', 'equipo', 'categoria', 'tiempo_inscripcion']], 
                use_container_width=True,
                hide_index=True
            )
        
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return
    
    st.markdown("---")
    
    # Tres columnas para las opciones de generación
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("### 📄 Papeletas PDF")
        st.info("Formato: Una papeleta por página en orientación horizontal")
        
        if st.button("🚀 Generar Papeletas PDF", type="primary", key="gen_pdf"):
            with st.spinner("Generando papeletas PDF..."):
                try:
                    success, message = papeletas_pdf_module.generar_papeletas_pdf()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                except Exception as e:
                    st.error(f"Error al generar papeletas PDF: {e}")
        
        # Descarga PDF
        if os.path.exists("papeletas_jueces.pdf"):
            with open("papeletas_jueces.pdf", "rb") as file:
                st.download_button(
                    label="⬇️ Descargar Papeletas PDF",
                    data=file.read(),
                    file_name="papeletas_jueces.pdf",
                    mime="application/pdf",
                    key="download_pdf"
                )
    
    with col2:
        st.markdown("### 📊 Papeletas Excel") 
        st.info("Formato: Rectángulos individuales por nadador, ideal para recortar")
        
        if st.button("🚀 Generar Papeletas Excel", type="primary", key="gen_excel"):
            with st.spinner("Generando papeletas Excel..."):
                try:
                    success, message = papeletas_excel_module.generar_papeletas_excel()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                except Exception as e:
                    st.error(f"Error al generar papeletas Excel: {e}")
        
        # Descarga Excel
        if os.path.exists("papeletas_jueces.xlsx"):
            with open("papeletas_jueces.xlsx", "rb") as file:
                st.download_button(
                    label="⬇️ Descargar Papeletas Excel",
                    data=file.read(),
                    file_name="papeletas_jueces.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel"
                )

    with col3:
        st.markdown("### 📊 Papeletas 3x3 Excel")
        st.info("Formato: Exacto como Excel, 3 papeletas por fila para imprimir y recortar")

        if st.button("🚀 Generar 3x3 Excel", type="primary", key="gen_excel_3x3"):
            with st.spinner("Generando papeletas 3x3 Excel..."):
                try:
                    success, message = papeletas_pdf_module.generar_papeletas_pdf_excel_3_per_row()
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                except Exception as e:
                    st.error(f"Error al generar papeletas 3x3 Excel: {e}")

        # Descarga 3x3 Excel PDF
        if os.path.exists("papeletas_jueces_excel_3_per_row.pdf"):
            with open("papeletas_jueces_excel_3_per_row.pdf", "rb") as file:
                st.download_button(
                    label="⬇️ Descargar 3x3 Excel PDF",
                    data=file.read(),
                    file_name="papeletas_jueces_excel_3_per_row.pdf",
                    mime="application/pdf",
                    key="download_excel_3x3"
                )

def gestion_archivos():
    st.markdown("## 📁 Gestión de Archivos")
    
    # Upload de archivos
    st.markdown("### ⬆️ Subir Archivos")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📋 Planilla de Inscripción")
        uploaded_inscripcion = st.file_uploader(
            "Sube tu archivo de inscripciones",
            type=['xlsx'],
            key="inscripcion",
            help="Archivo Excel con datos de nadadores registrados"
        )
        
        if uploaded_inscripcion:
            with open("planilla_inscripcion.xlsx", "wb") as f:
                f.write(uploaded_inscripcion.getbuffer())
            st.success("✅ Archivo de inscripción subido correctamente")
    
    with col2:
        st.markdown("#### 🏆 Resultados de Competencia")
        uploaded_resultados = st.file_uploader(
            "Sube el archivo de resultados",
            type=['xlsx'],
            key="resultados",
            help="Archivo Excel con tiempos finales de la competencia"
        )
        
        if uploaded_resultados:
            with open("resultados_con_tiempos.xlsx", "wb") as f:
                f.write(uploaded_resultados.getbuffer())
            st.success("✅ Archivo de resultados subido correctamente")
    
    # Nueva sección para base de datos
    st.markdown("### 💾 Base de Datos de Atletas")
    
    col3, col4 = st.columns(2)
    
    with col3:
        st.markdown("#### 📊 Base de Datos Local")
        if os.path.exists("BASE-DE-DATOS.xlsx"):
            st.success("✅ Base de datos local disponible")
            # Mostrar información de la base de datos local
            try:
                xl_file = pd.ExcelFile("BASE-DE-DATOS.xlsx")
                target_sheets = ['FPROYECCION 2025T', 'M. PROYECCION 2025']
                available_sheets = [s for s in target_sheets if s in xl_file.sheet_names]
                st.info(f"🔍 Hojas disponibles: {', '.join(available_sheets)}")
                
                # Contar registros totales
                total_records = 0
                for sheet in available_sheets:
                    try:
                        df = pd.read_excel("BASE-DE-DATOS.xlsx", sheet_name=sheet)
                        total_records += len(df)
                    except:
                        pass
                st.info(f"📈 Total de registros: {total_records:,}")
                
            except Exception as e:
                pass  # No mostrar error innecesario
        else:
            pass  # No mostrar warning innecesario si no hay base de datos
    
    with col4:
        st.markdown("#### 🔄 Cargar Base de Datos Externa")
        uploaded_database = st.file_uploader(
            "Sube tu propia base de datos",
            type=['xlsx'],
            key="database",
            help="Archivo Excel con base de datos de atletas (debe contener hojas FPROYECCION 2025T y M. PROYECCION 2025)"
        )
        
        if uploaded_database:
            try:
                # Guardar como archivo temporal y verificar estructura
                with open("BASE-DE-DATOS-TEMP.xlsx", "wb") as f:
                    f.write(uploaded_database.getbuffer())
                
                # Verificar estructura
                xl_file = pd.ExcelFile("BASE-DE-DATOS-TEMP.xlsx")
                target_sheets = ['FPROYECCION 2025T', 'M. PROYECCION 2025']
                available_sheets = [s for s in target_sheets if s in xl_file.sheet_names]
                
                if available_sheets:
                    # Reemplazar base de datos actual
                    if os.path.exists("BASE-DE-DATOS.xlsx"):
                        os.rename("BASE-DE-DATOS.xlsx", "BASE-DE-DATOS-BACKUP.xlsx")
                    os.rename("BASE-DE-DATOS-TEMP.xlsx", "BASE-DE-DATOS.xlsx")
                    
                    st.success("✅ Base de datos externa cargada correctamente")
                    st.info(f"🔍 Hojas encontradas: {', '.join(available_sheets)}")
                    
                    # Contar registros
                    total_records = 0
                    for sheet in available_sheets:
                        try:
                            df = pd.read_excel("BASE-DE-DATOS.xlsx", sheet_name=sheet)
                            total_records += len(df)
                        except:
                            pass
                    st.info(f"📈 Total de registros cargados: {total_records:,}")
                    st.rerun()
                else:
                    os.remove("BASE-DE-DATOS-TEMP.xlsx")
                    st.error("❌ El archivo no contiene las hojas requeridas (FPROYECCION 2025T o M. PROYECCION 2025)")
                    
            except Exception as e:
                st.error(f"❌ Error al procesar la base de datos: {e}")
                if os.path.exists("BASE-DE-DATOS-TEMP.xlsx"):
                    os.remove("BASE-DE-DATOS-TEMP.xlsx")
    
    # Opción para restaurar base de datos original
    if os.path.exists("BASE-DE-DATOS-BACKUP.xlsx"):
        st.markdown("#### 🔄 Restaurar Base de Datos Original")
        col5, col6 = st.columns([2, 1])
        
        with col5:
            st.info("📁 Se encontró un respaldo de la base de datos original")
        
        with col6:
            if st.button("♻️ Restaurar Original", help="Restaurar la base de datos original del repositorio"):
                try:
                    if os.path.exists("BASE-DE-DATOS.xlsx"):
                        os.remove("BASE-DE-DATOS.xlsx")
                    os.rename("BASE-DE-DATOS-BACKUP.xlsx", "BASE-DE-DATOS.xlsx")
                    st.success("✅ Base de datos original restaurada")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error al restaurar: {e}")
    
    # Mostrar archivos existentes
    st.markdown("### 📄 Archivos Disponibles")
    
    archivos = {
        "planilla_inscripcion.xlsx": "📋 Planilla de Inscripción",
        "BASE-DE-DATOS.xlsx": "🗄️ Base de Datos de Atletas",
        "sembrado_competencia.xlsx": "📊 Sembrado por Categoría",
        "sembrado_competencia_POR_TIEMPO.xlsx": "⏱️ Sembrado por Tiempo",
        "resultados_con_tiempos.xlsx": "🏆 Resultados de Competencia",
        "reporte_premiacion_final_CORREGIDO.xlsx": "🏅 Reporte de Premiación"
    }
    
    # Detectar archivos dinámicos generados por el procesador de sembrado
    import glob
    dynamic_result_files = glob.glob("resultados_desde_sembrado_*.xlsx")
    for file in dynamic_result_files:
        archivos[file] = "🏁 Resultados desde Sembrado"
    
    for archivo, descripcion in archivos.items():
        if os.path.exists(archivo):
            col1, col2, col3 = st.columns([3, 1, 1])
            
            with col1:
                st.write(f"{descripcion}")
            
            with col2:
                try:
                    df = pd.read_excel(archivo, nrows=10)
                    if st.button("👁️", key=f"view_{archivo}", help="Ver vista previa"):
                        st.dataframe(df)
                except:
                    st.write("📄")
            
            with col3:
                with open(archivo, "rb") as file:
                    st.download_button(
                        label="⬇️",
                        data=file.read(),
                        file_name=archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{archivo}",
                        help="Descargar archivo"
                    )

def inscripcion_nadadores_interface():
    st.markdown("## ✍️ Inscripción de Nadadores")
    
    # Inicializar el sistema de inscripción
    registration_system = inscripcion_nadadores.SwimmerRegistration()
    
    # Tabs para diferentes funciones
    tab1, tab2, tab3 = st.tabs(["➕ Nuevo Nadador", "📝 Nadadores Inscritos", "📊 Reporte de Inscripción"])
    
    with tab1:
        st.markdown("### Registrar Nuevo Nadador")
        
        # Método de inscripción
        inscripcion_method = st.radio(
            "Método de inscripción:",
            ["✍️ Manual", "🔍 Buscar en Base de Datos", "📤 Importar desde Excel"],
            horizontal=True
        )
        
        if inscripcion_method == "✍️ Manual":
            # INSCRIPCIÓN MANUAL (código existente)
            col1, col2 = st.columns(2)
            
            with col1:
                name = st.text_input("Nombre y Apellidos", placeholder="Ej: Juan Pérez García")
                team = st.text_input("Equipo", placeholder="Ej: Club Natación TEN")
                from datetime import datetime, date
                birth_date = st.date_input("Fecha de Nacimiento",
                                         value=None,
                                         min_value=date(1900, 1, 1),
                                         max_value=date.today(),
                                         help="Selecciona la fecha de nacimiento del nadador")

            with col2:
                gender = st.selectbox("Sexo", ["M", "F"], format_func=lambda x: "Masculino" if x == "M" else "Femenino")

                if birth_date:
                    age, category = registration_system.resolve_swimmer_age_and_category(gender, birth_date=birth_date)
                    reference_date, _ = registration_system.get_event_age_reference()
                    st.info(f"Edad al {reference_date.strftime('%d/%m/%Y')}: **{age} años**")
                    st.info(f"Categoría automática: **{category}**")
                else:
                    age = None
                    category = None
                    st.warning("Selecciona la fecha de nacimiento para calcular edad y categoría")
                
            st.markdown("### Pruebas de Inscripción")
            st.markdown("*Opcional: ingresa tiempos (MM:SS.dd). Las pruebas sin tiempo se inscriben automáticamente como **s/t**. Se inscribe en **todas** las pruebas de la categoría.*")
            
            events_data = {}
            col1, col2, col3 = st.columns(3)
            
            # Obtener eventos disponibles filtrados por edad para la categoría del nadador
            available_events = registration_system.get_available_events_for_swimmer_category(category, age) if category else []

            # Mostrar información sobre restricciones de edad si hay eventos filtrados
            all_events = registration_system.get_available_events()
            if len(available_events) < len(all_events):
                excluded_events = [e for e in all_events if e not in available_events]
                if excluded_events:
                    st.info(f"ℹ️ Eventos no disponibles para edad {age}: {', '.join(excluded_events)}")

            for i, event in enumerate(available_events):
                with [col1, col2, col3][i % 3]:
                    time_input = st.text_input(
                        event,
                        key=f"manual_event_{i}",
                        placeholder="MM:SS.dd",
                        help="Ejemplo: 1:25.30 o 85.30"
                    )
                    
                    if time_input:
                        is_valid, error_msg = registration_system.validate_time_format(time_input)
                        if not is_valid:
                            st.error(error_msg)
                        else:
                            events_data[event] = time_input
            
            if st.button("🏊‍♂️ Registrar Nadador (Manual)", type="primary"):
                if not name.strip():
                    st.error("El nombre es obligatorio")
                elif not team.strip():
                    st.error("El equipo es obligatorio")
                elif not birth_date:
                    st.error("La fecha de nacimiento es obligatoria")
                elif age is None or not category:
                    st.error("No se pudo calcular la edad o categoría")
                else:
                    events_data = registration_system.complete_category_events(category, age, events_data)
                    swimmer_data = {
                        'name': name.strip(),
                        'team': team.strip(),
                        'age': age,
                        'birth_date': birth_date,
                        'category': category,
                        'gender': gender,
                        'events': events_data
                    }
                    
                    success, message, duplicate_info = registration_system.add_swimmer(swimmer_data)
                    if success:
                        st.success(message)
                        st.balloons()
                        st.rerun()
                    else:
                        if duplicate_info:  # Es un duplicado
                            st.error(message)
                            st.markdown("### ⚠️ Información del Nadador Existente:")
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**Nombre:** {duplicate_info['name']}")
                                st.write(f"**Equipo:** {duplicate_info['team']}")
                                st.write(f"**Edad:** {duplicate_info['age']}")
                            with col2:
                                st.write(f"**Categoría:** {duplicate_info['category']}")
                                st.write(f"**Sexo:** {'Masculino' if duplicate_info['gender'] == 'M' else 'Femenino'}")
                            
                            if st.button("🚫 Inscribir De Todas Formas", key="force_add_manual"):
                                success_force, message_force, _ = registration_system.add_swimmer(swimmer_data, force_add=True)
                                if success_force:
                                    st.success(f"✅ {swimmer_data['name']} inscrito como registro adicional")
                                    st.balloons()
                                    st.rerun()
                                else:
                                    st.error(message_force)
                        else:
                            st.error(message)
        
        elif inscripcion_method == "🔍 Buscar en Base de Datos":
            # BÚSQUEDA EN BASE DE DATOS
            st.markdown("### Buscar Nadador en Base de Datos")
            
            # Estado de la base de datos
            with st.container():
                col_db1, col_db2 = st.columns([2, 1])
                
                with col_db1:
                    if os.path.exists("BASE-DE-DATOS.xlsx"):
                        try:
                            xl_file = pd.ExcelFile("BASE-DE-DATOS.xlsx")
                            target_sheets = ['FPROYECCION 2025T', 'M. PROYECCION 2025']
                            available_sheets = [s for s in target_sheets if s in xl_file.sheet_names]
                            
                            if available_sheets:
                                # Contar registros
                                total_records = 0
                                for sheet in available_sheets:
                                    try:
                                        df = pd.read_excel("BASE-DE-DATOS.xlsx", sheet_name=sheet)
                                        total_records += len(df)
                                    except:
                                        pass
                                
                                st.success(f"✅ Base de datos activa: {total_records:,} atletas ({len(available_sheets)} hojas)")
                            else:
                                st.warning("⚠️ Base de datos sin hojas válidas")
                        except Exception as e:
                            pass  # No mostrar error innecesario
                    else:
                        pass  # No mostrar error innecesario si no hay base de datos
                
                with col_db2:
                    with st.popover("🔄 Cambiar BD", help="Cargar una base de datos diferente"):
                        st.markdown("**Cargar nueva base de datos:**")
                        uploaded_db = st.file_uploader(
                            "Archivo de base de datos",
                            type=['xlsx'],
                            key="db_for_search",
                            help="Debe contener hojas FPROYECCION 2025T y/o M. PROYECCION 2025"
                        )
                        
                        if uploaded_db:
                            try:
                                # Guardar como archivo temporal
                                with open("BASE-DE-DATOS-TEMP.xlsx", "wb") as f:
                                    f.write(uploaded_db.getbuffer())
                                
                                # Verificar estructura
                                xl_file = pd.ExcelFile("BASE-DE-DATOS-TEMP.xlsx")
                                target_sheets = ['FPROYECCION 2025T', 'M. PROYECCION 2025']
                                available_sheets = [s for s in target_sheets if s in xl_file.sheet_names]
                                
                                if available_sheets:
                                    # Crear backup si existe BD actual
                                    if os.path.exists("BASE-DE-DATOS.xlsx"):
                                        os.rename("BASE-DE-DATOS.xlsx", "BASE-DE-DATOS-BACKUP.xlsx")
                                    os.rename("BASE-DE-DATOS-TEMP.xlsx", "BASE-DE-DATOS.xlsx")
                                    
                                    st.success("✅ Base de datos cargada")
                                    st.rerun()
                                else:
                                    os.remove("BASE-DE-DATOS-TEMP.xlsx")
                                    st.error("❌ Archivo sin hojas válidas")
                            except Exception as e:
                                st.error(f"❌ Error: {e}")
                                if os.path.exists("BASE-DE-DATOS-TEMP.xlsx"):
                                    os.remove("BASE-DE-DATOS-TEMP.xlsx")
            
            st.markdown("---")
            
            search_term = st.text_input(
                "Buscar nadador por nombre:",
                placeholder="Escribe el nombre del nadador...",
                help="Se buscará en la base de datos activa"
            )
            
            if search_term and len(search_term.strip()) >= 3:
                with st.spinner("Buscando en la base de datos..."):
                    matches, search_message = registration_system.search_swimmer_in_database(search_term)
                
                if matches:
                    st.success(search_message)
                    
                    # Mostrar resultados de búsqueda
                    st.markdown("### Resultados de Búsqueda")
                    
                    for i, match in enumerate(matches):
                        with st.expander(f"🏊‍♂️ {match['name']}"):
                            col1, col2 = st.columns([3, 1])
                            
                            with col1:
                                # Obtener información del nadador
                                swimmer_info = registration_system.get_swimmer_info_from_database(match)
                                st.write(f"**Equipo:** {swimmer_info['team'] or 'No especificado'}")
                                st.write(f"**Edad:** {swimmer_info['age'] or 'No especificada'}")
                                st.write(f"**Categoría:** {swimmer_info['category'] or 'No especificada'}")
                                st.write(f"**Sexo:** {'Masculino' if swimmer_info['gender'] == 'M' else 'Femenino' if swimmer_info['gender'] == 'F' else 'No especificado'}")
                                
                                # Mostrar tiempos disponibles
                                latest_times, times_message = registration_system.get_swimmer_latest_times(swimmer_info)
                                if latest_times:
                                    st.write("**Últimos tiempos registrados:**")
                                    for event, time in latest_times.items():
                                        st.write(f"• {event}: {time}")
                                else:
                                    st.write("*No hay tiempos registrados*")
                            
                            with col2:
                                if st.button(f"📋 Inscribir", key=f"db_register_{i}"):
                                    # Crear nadador desde base de datos
                                    swimmer_data, create_message = registration_system.create_swimmer_from_database(match)
                                    
                                    # Registrar el nadador
                                    success, register_message, duplicate_info = registration_system.add_swimmer(swimmer_data)
                                    if success:
                                        st.success(register_message)
                                        st.balloons()
                                        st.rerun()
                                    else:
                                        if duplicate_info:  # Es un duplicado
                                            st.error(register_message)
                                            if st.button("🚫 Inscribir De Todas Formas", key=f"force_add_db_{i}"):
                                                success_force, message_force, _ = registration_system.add_swimmer(swimmer_data, force_add=True)
                                                if success_force:
                                                    st.success(f"✅ {swimmer_data['name']} inscrito como registro adicional")
                                                    st.balloons()
                                                    st.rerun()
                                                else:
                                                    st.error(message_force)
                                        else:
                                            st.error(register_message)
                else:
                    st.warning(search_message)
            
            elif search_term and len(search_term.strip()) < 3:
                st.info("Escribe al menos 3 caracteres para buscar")
            
            # Información sobre la base de datos
            if os.path.exists(registration_system.archivo_base_datos):
                st.markdown("### 📊 Estado de la Base de Datos")
                st.success(f"✅ Base de datos encontrada: `{registration_system.archivo_base_datos}`")
            else:
                st.warning(f"⚠️ No se encontró la base de datos: `{registration_system.archivo_base_datos}`")
        
        elif inscripcion_method == "📤 Importar desde Excel":
            # IMPORTACIÓN MASIVA DESDE EXCEL
            st.markdown("### Importar Nadadores desde Excel")
            
            st.markdown("""
            <div class="info-message">
                <h4>📋 Instrucciones para la importación masiva:</h4>
                <ol>
                    <li>Utiliza un archivo Excel con la misma estructura que el archivo <code>planilla_inscripcion.xlsx</code></li>
                    <li>Las columnas requeridas son: <strong>NOMBRE Y AP</strong>, <strong>EQUIPO</strong>, <strong>EDAD</strong>, <strong>CAT.</strong>, <strong>SEXO</strong></li>
                    <li>Las columnas de pruebas pueden usar texto <strong>MM:SS,dd</strong> o <strong>MM:SS.dd</strong> (ej.: 1:25,30), celda de hora de Excel, o segundos totales como número</li>
                    <li>Los nadadores duplicados serán omitidos automáticamente</li>
                    <li>La categoría se calculará automáticamente si no se proporciona</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)
            
            # Descarga del template
            col1, col2 = st.columns([2, 1])
            with col1:
                st.markdown("#### 📥 Subir Archivo de Inscripciones")
            
            with col2:
                if os.path.exists("planilla_inscripcion.xlsx"):
                    with open("planilla_inscripcion.xlsx", "rb") as file:
                        st.download_button(
                            label="📄 Descargar Template",
                            data=file.read(),
                            file_name="template_inscripciones.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Descarga el archivo actual como plantilla"
                        )
            
            # Upload del archivo
            uploaded_file = st.file_uploader(
                "Selecciona el archivo Excel con los nadadores:",
                type=['xlsx', 'xls'],
                help="El archivo debe contener las columnas requeridas y los datos de los nadadores",
                key="bulk_import_uploader"
            )
            
            # Limpiar estado cuando se cambia de archivo
            if uploaded_file is not None:
                # Detectar si es un archivo nuevo
                current_file_info = f"{uploaded_file.name}_{uploaded_file.size}"
                if 'last_uploaded_file' not in st.session_state:
                    st.session_state.last_uploaded_file = current_file_info
                elif st.session_state.last_uploaded_file != current_file_info:
                    # Archivo nuevo detectado, limpiar estados previos
                    st.session_state.last_uploaded_file = current_file_info
                    # Limpiar cualquier resultado anterior
                    for key in list(st.session_state.keys()):
                        if key.startswith('bulk_import_'):
                            del st.session_state[key]
                    st.rerun()
                
                st.success(f"✅ Archivo cargado: {uploaded_file.name}")
                
                # Vista previa del archivo
                if st.checkbox("👁️ Ver vista previa del archivo", key="preview_checkbox"):
                    try:
                        # Usar caché para evitar re-lecturas
                        @st.cache_data
                        def load_preview_data(file_content, file_name):
                            return pd.read_excel(io.BytesIO(file_content))
                        
                        preview_df = load_preview_data(uploaded_file.getvalue(), uploaded_file.name)
                        st.markdown("**Vista previa (primeras 10 filas):**")
                        st.dataframe(preview_df.head(10))
                        st.info(f"Total de filas en el archivo: {len(preview_df)}")
                    except Exception as e:
                        st.error(f"Error al leer el archivo: {str(e)}")
                
                # Botón de importación
                st.markdown("---")
                col1, col2, col3 = st.columns([1, 1, 2])
                
                with col2:
                    if st.button("🚀 Importar Nadadores", type="primary", use_container_width=True):
                        with st.spinner("Importando nadadores..."):
                            success, message = registration_system.bulk_import_from_excel(uploaded_file)
                            
                            # Guardar resultado en session_state
                            st.session_state.bulk_import_result = (success, message)
                            st.session_state.bulk_import_completed = True
                            
                            if success:
                                st.success("✅ **Importación completada exitosamente!**")
                                st.markdown(message)
                                st.balloons()
                                
                                # Mostrar estadísticas actualizadas
                                updated_swimmers = registration_system.get_swimmers_list()
                                if updated_swimmers:
                                    st.markdown("---")
                                    st.markdown("### 📊 Estado Actualizado de Inscripciones")
                                    
                                    total_swimmers = len(updated_swimmers)
                                    male_count = len([s for s in updated_swimmers if s['gender'] == 'M'])
                                    female_count = len([s for s in updated_swimmers if s['gender'] == 'F'])
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Total Nadadores", total_swimmers)
                                    with col2:
                                        st.metric("Masculinos", male_count)
                                    with col3:
                                        st.metric("Femeninos", female_count)
                            else:
                                st.error("❌ **Error en la importación**")
                                st.markdown(message)
                
                # Mostrar resultados previos si existen
                if 'bulk_import_completed' in st.session_state and st.session_state.bulk_import_completed:
                    st.markdown("---")
                    col1, col2 = st.columns([3, 1])
                    
                    with col1:
                        if st.session_state.bulk_import_result[0]:
                            st.info("📋 **Importación anterior completada**")
                        else:
                            st.warning("⚠️ **Última importación tuvo errores**")
                    
                    with col2:
                        if st.button("🧹 Limpiar Resultados", help="Limpia los resultados de la importación anterior"):
                            # Limpiar resultados
                            for key in list(st.session_state.keys()):
                                if key.startswith('bulk_import_'):
                                    del st.session_state[key]
                            st.rerun()
            
            else:
                st.info("📤 Selecciona un archivo Excel para importar nadadores masivamente")
    
    with tab2:
        st.markdown("### Nadadores Inscritos")
        
        swimmers = registration_system.get_swimmers_list()
        
        if not swimmers:
            st.info("No hay nadadores inscritos aún. Usa la pestaña 'Nuevo Nadador' para registrar.")
        else:
            st.success(f"Total de nadadores inscritos: **{len(swimmers)}**")
            
            for i, swimmer in enumerate(swimmers):
                with st.expander(f"🏊‍♂️ {swimmer['name']} - {swimmer['team']} ({swimmer['age']} años)"):
                    col1, col2 = st.columns([3, 1])
                    
                    with col1:
                        st.write(f"**Categoría:** {swimmer['category']}")
                        st.write(f"**Sexo:** {'Masculino' if swimmer['gender'] == 'M' else 'Femenino'}")
                        
                        if swimmer['events']:
                            st.write("**Pruebas inscritas:**")
                            for event in swimmer['events']:
                                st.write(f"• {event}")
                        else:
                            st.write("*Sin pruebas registradas*")
                    
                    with col2:
                        col_edit, col_delete = st.columns(2)
                        
                        with col_edit:
                            if st.button("✏️ Editar", key=f"edit_{i}"):
                                st.session_state[f'editing_swimmer_{i}'] = True
                                st.rerun()
                        
                        with col_delete:
                            if st.button("🗑️ Eliminar", key=f"delete_{i}"):
                                success, message = registration_system.delete_swimmer(swimmer['index'])
                                if success:
                                    st.success(message)
                                    st.rerun()
                                else:
                                    st.error(message)
            
            # Formulario de edición (aparece cuando se hace clic en editar)
            for i, swimmer in enumerate(swimmers):
                if f'editing_swimmer_{i}' in st.session_state and st.session_state[f'editing_swimmer_{i}']:
                    st.markdown("---")
                    st.markdown(f"### ✏️ Editando: {swimmer['name']}")
                    
                    # Obtener datos actuales del nadador
                    swimmer_data, message = registration_system.get_swimmer_for_editing(swimmer['index'])
                    
                    if swimmer_data:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            edit_name = st.text_input("Nombre y Apellidos", value=swimmer_data['name'], key=f"edit_name_{i}")
                            edit_team = st.text_input("Equipo", value=swimmer_data['team'], key=f"edit_team_{i}")

                            # Usar fecha de nacimiento si está disponible, sino calcular desde la edad
                            current_birth_date = swimmer_data.get('birth_date')
                            if current_birth_date is None:
                                # Si no hay fecha de nacimiento, estimar desde la edad
                                current_birth_date = date.today().replace(year=date.today().year - swimmer_data['age'])
                            elif isinstance(current_birth_date, str):
                                try:
                                    current_birth_date = datetime.strptime(current_birth_date, '%Y-%m-%d').date()
                                except:
                                    current_birth_date = date.today().replace(year=date.today().year - swimmer_data['age'])

                            edit_birth_date = st.date_input("Fecha de Nacimiento",
                                                           value=current_birth_date,
                                                           min_value=date(1900, 1, 1),
                                                           max_value=date.today(),
                                                           key=f"edit_birth_date_{i}",
                                                           help="Selecciona la fecha de nacimiento del nadador")

                        with col2:
                            edit_gender = st.selectbox("Sexo", ["M", "F"],
                                                     index=0 if swimmer_data['gender'] == 'M' else 1,
                                                     format_func=lambda x: "Masculino" if x == "M" else "Femenino",
                                                     key=f"edit_gender_{i}")

                            if edit_birth_date:
                                edit_age, edit_category = registration_system.resolve_swimmer_age_and_category(
                                    edit_gender, birth_date=edit_birth_date
                                )
                                reference_date, _ = registration_system.get_event_age_reference()
                                st.info(f"Edad al {reference_date.strftime('%d/%m/%Y')}: **{edit_age} años**")
                                st.info(f"Categoría automática: **{edit_category}**")
                            else:
                                edit_age = swimmer_data['age']
                                edit_category = registration_system.get_category_by_age(
                                    edit_age, edit_gender, swimmer_data.get('birth_date')
                                )
                                st.info(f"Categoría automática: **{edit_category}**")
                        
                        st.markdown("### Pruebas de Inscripción")
                        st.markdown("*Edita los tiempos de inscripción. Deja en blanco para eliminar la prueba.*")
                        
                        edit_events_data = {}
                        col1, col2, col3 = st.columns(3)

                        # Obtener eventos disponibles filtrados por edad para la categoría del nadador
                        edit_available_events = registration_system.get_available_events_for_swimmer_category(edit_category, edit_age)

                        # Mostrar información sobre restricciones de edad si hay eventos filtrados
                        all_edit_events = registration_system.get_available_events()
                        if len(edit_available_events) < len(all_edit_events):
                            excluded_edit_events = [e for e in all_edit_events if e not in edit_available_events]
                            if excluded_edit_events:
                                st.info(f"ℹ️ Eventos no disponibles para edad {edit_age}: {', '.join(excluded_edit_events)}")

                        for j, event in enumerate(edit_available_events):
                            with [col1, col2, col3][j % 3]:
                                current_time = swimmer_data['events'].get(event, "")
                                edit_time_input = st.text_input(
                                    event,
                                    value=current_time,
                                    key=f"edit_event_{i}_{j}",
                                    placeholder="MM:SS.dd",
                                    help="Ejemplo: 1:25.30 o 85.30"
                                )
                                
                                if edit_time_input:
                                    is_valid, error_msg = registration_system.validate_time_format(edit_time_input)
                                    if not is_valid:
                                        st.error(error_msg)
                                    else:
                                        edit_events_data[event] = edit_time_input
                        
                        col_save, col_cancel = st.columns(2)
                        
                        with col_save:
                            if st.button("💾 Guardar Cambios", key=f"save_{i}", type="primary"):
                                updated_swimmer_data = {
                                    'name': edit_name.strip(),
                                    'team': edit_team.strip(),
                                    'age': edit_age,
                                    'birth_date': edit_birth_date,
                                    'category': edit_category,
                                    'gender': edit_gender,
                                    'events': edit_events_data
                                }
                                
                                success, message = registration_system.update_swimmer(swimmer['index'], updated_swimmer_data)
                                if success:
                                    st.success(message)
                                    del st.session_state[f'editing_swimmer_{i}']
                                    st.balloons()
                                    st.rerun()
                                else:
                                    st.error(message)
                        
                        with col_cancel:
                            if st.button("❌ Cancelar", key=f"cancel_{i}"):
                                del st.session_state[f'editing_swimmer_{i}']
                                st.rerun()
                    else:
                        st.error(message)
    
    with tab3:
        st.markdown("### 📊 Reporte de Inscripción")

        # Datos y agregados para PDF y estadísticas (debe existir antes del botón PDF)
        swimmers = registration_system.get_swimmers_list()
        teams = {}
        categories = {}
        genders = {"Masculino": 0, "Femenino": 0}
        events_stats = {}
        for swimmer in swimmers:
            teams[swimmer['team']] = teams.get(swimmer['team'], 0) + 1
            categories[swimmer['category']] = categories.get(swimmer['category'], 0) + 1
            gender_label = "Masculino" if swimmer['gender'] == 'M' else "Femenino"
            genders[gender_label] += 1
            for event_info in swimmer['events']:
                if ':' in event_info:
                    event_name = event_info.split(':')[0].strip()
                    events_stats[event_name] = events_stats.get(event_name, 0) + 1

        # Sección de Acciones del Sistema (movida desde pagos)
        st.markdown("#### 🔧 Acciones del Sistema")

        col_actions1, col_actions2, col_actions3 = st.columns(3)

        with col_actions1:
            if st.button("📋 Crear Archivo Vacío"):
                success, message = registration_system.create_empty_registration_file()
                if success:
                    st.success(message)
                else:
                    st.error(message)

        with col_actions2:
            if os.path.exists(registration_system.archivo_inscripcion):
                with open(registration_system.archivo_inscripcion, "rb") as file:
                    st.download_button(
                        label="📥 Descargar Planilla de Inscripción",
                        data=file.read(),
                        file_name="planilla_inscripcion.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        with col_actions3:
            # Generar reporte PDF (bytes en session_state para que el download_button persista entre reruns)
            _pdf_key = "_reporte_inscripcion_pdf_bytes"
            if swimmers:
                if st.button("📄 Generar Reporte PDF", type="primary"):
                    try:
                        pdf_data = registration_system.generate_pdf_report(swimmers, teams, categories, genders, events_stats)
                        if pdf_data:
                            st.session_state[_pdf_key] = pdf_data
                            st.success("¡Reporte PDF listo! Usa **Descargar** justo debajo.")
                        else:
                            st.session_state.pop(_pdf_key, None)
                            st.error("**ReportLab no está instalado en este entorno Python**")
                            st.markdown("""
                            **Para generar reportes PDF, pruebe uno de estos comandos:**
                            ```bash
                            pip install reportlab
                            # o
                            pip3 install reportlab
                            # o
                            python -m pip install reportlab
                            # o
                            python3 -m pip install reportlab
                            ```

                            **Si está usando un entorno virtual, asegúrese de activarlo primero:**
                            ```bash
                            source venv/bin/activate  # Linux/Mac
                            # o
                            venv\\Scripts\\activate     # Windows
                            ```

                            Luego reinicie la aplicación Streamlit.
                            """)
                    except Exception as e:
                        st.session_state.pop(_pdf_key, None)
                        if "reportlab" in str(e).lower() or "not available" in str(e).lower():
                            st.error("Para generar reportes PDF, instale la librería ReportLab ejecutando: pip install reportlab")
                        else:
                            st.error(f"Error al generar el reporte PDF: {str(e)}")

                if st.session_state.get(_pdf_key):
                    st.download_button(
                        label="📄 Descargar Reporte PDF",
                        data=st.session_state[_pdf_key],
                        file_name="reporte_inscripciones.pdf",
                        mime="application/pdf",
                        key="download_reporte_inscripciones_pdf",
                    )
            else:
                st.session_state.pop(_pdf_key, None)
                st.info("No hay nadadores inscritos para generar reporte")

        # Sección de limpieza de inscripciones
        st.markdown("#### 🧹 Limpiar Inscripciones")
        col_clean1, col_clean2 = st.columns([2, 1])

        with col_clean1:
            st.info("🗑️ Eliminar todas las inscripciones para empezar con nuevos nadadores")

        with col_clean2:
            if st.button("👥 Limpiar Inscripciones", type="secondary", help="Eliminar planilla de inscripción"):
                if os.path.exists("planilla_inscripcion.xlsx"):
                    try:
                        os.remove("planilla_inscripcion.xlsx")
                        st.success("✅ Inscripciones eliminadas")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al limpiar: {e}")
                else:
                    st.info("No hay archivo de inscripciones para limpiar")

        st.markdown("---")

        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📊 Estadísticas")
            
            if swimmers:
                # Métricas principales
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Nadadores", len(swimmers))
                with col2:
                    st.metric("Equipos Diferentes", len(teams))
                with col3:
                    st.metric("Categorías", len(categories))
                
                # Distribución por género
                st.markdown("#### 👥 Distribución por Sexo")
                if genders:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("👨 Masculino", genders["Masculino"])
                    with col2:
                        st.metric("👩 Femenino", genders["Femenino"])
                    
                    # Gráfico de género
                    st.bar_chart(genders)
                
                # Distribución por pruebas
                if events_stats:
                    st.markdown("#### 🏊‍♂️ Distribución por Pruebas")
                    st.bar_chart(events_stats)
                    
                    # Top 5 pruebas más populares
                    top_events = sorted(events_stats.items(), key=lambda x: x[1], reverse=True)[:5]
                    st.markdown("**Top 5 Pruebas Más Populares:**")
                    for i, (event, count) in enumerate(top_events, 1):
                        st.write(f"{i}. **{event}**: {count} nadadores")
                
                # Otras estadísticas expandibles
                if st.checkbox("📊 Ver distribución por categorías"):
                    st.bar_chart(categories)
                
                if st.checkbox("🏢 Ver distribución por equipos"):
                    st.bar_chart(teams)
                
                if st.checkbox("📈 Estadísticas detalladas"):
                    st.markdown("#### Estadísticas Detalladas")
                    
                    # Promedio de edad
                    ages = [swimmer['age'] for swimmer in swimmers if swimmer['age']]
                    if ages:
                        avg_age = sum(ages) / len(ages)
                        st.metric("Edad Promedio", f"{avg_age:.1f} años")
                    
                    # Nadadores por categoría
                    st.markdown("**Nadadores por Categoría:**")
                    for category, count in sorted(categories.items()):
                        percentage = (count / len(swimmers)) * 100
                        st.write(f"• **{category}**: {count} nadadores ({percentage:.1f}%)")
                    
                    # Total de inscripciones en pruebas
                    total_event_entries = sum(len(swimmer['events']) for swimmer in swimmers)
                    if total_event_entries > 0:
                        avg_events_per_swimmer = total_event_entries / len(swimmers)
                        st.metric("Promedio pruebas/nadador", f"{avg_events_per_swimmer:.1f}")
            else:
                st.info("No hay datos para mostrar estadísticas")

        # Sección de Reporte por Equipos
        if swimmers:
            st.markdown("---")
            st.markdown("### 🏊‍♂️ Reporte por Equipos")

            col1, col2 = st.columns(2)
            with col1:
                if st.button("👁️ Previsualizar Reporte por Equipos"):
                    with st.spinner("Generando previsualización..."):
                        teams_data, message = registration_system.preview_team_report()
                        if teams_data:
                            st.session_state['teams_preview'] = teams_data
                            st.success("Previsualización generada")
                        else:
                            st.error(message)

            with col2:
                if st.button("📊 Generar Reporte por Equipos"):
                    with st.spinner("Generando reporte por equipos..."):
                        teams_data, message = registration_system.generate_team_report()
                        if teams_data:
                            st.success(message)
                            st.session_state['teams_preview'] = teams_data
                            st.info("¡Reporte generado! Desplázate hacia abajo para descargar el Excel.")
                        else:
                            st.error(message)

            # Mostrar previsualización si está disponible
            if 'teams_preview' in st.session_state and st.session_state['teams_preview']:
                teams_data = st.session_state['teams_preview']

                # Mostrar resumen de equipos
                st.markdown("#### 📋 Resumen de Equipos")
                summary_data = []
                for team_name, team_info in teams_data.items():
                    summary_data.append({
                        'Equipo': team_name,
                        'Nadadores': team_info['total_swimmers'],
                        'Masculino': team_info['genders']['M'],
                        'Femenino': team_info['genders']['F'],
                        'Inscripciones': team_info['total_events']
                    })

                import pandas as pd
                df_summary = pd.DataFrame(summary_data)
                st.dataframe(df_summary, use_container_width=True)

                # Selector para generar PDF individual por equipo
                st.markdown("#### 📄 Generar PDF por Equipo Individual")

                # Crear lista de equipos para el selectbox
                team_names = list(teams_data.keys())
                if team_names:
                    selected_team = st.selectbox(
                        "Selecciona el equipo:",
                        team_names,
                        key="team_selector"
                    )

                    col_pdf1, col_pdf2 = st.columns(2)
                    with col_pdf1:
                        if st.button(f"📋 Generar PDF de {selected_team}", type="secondary"):
                            with st.spinner(f"Generando PDF para {selected_team}..."):
                                team_info = teams_data[selected_team]
                                pdf_data, filename = registration_system.generate_team_pdf(selected_team, team_info)

                                if pdf_data:
                                    st.success(f"✅ PDF generado para {selected_team}")
                                    st.download_button(
                                        label=f"⬇️ Descargar PDF de {selected_team}",
                                        data=pdf_data,
                                        file_name=filename,
                                        mime="application/pdf"
                                    )
                                else:
                                    st.error(f"❌ {filename}")

                    with col_pdf2:
                        # Mostrar estadísticas del equipo seleccionado
                        team_info = teams_data[selected_team]
                        st.info(
                            f"**{selected_team}**\n\n"
                            f"👥 {team_info['total_swimmers']} nadadores\n\n"
                            f"🏊‍♀️ {team_info['total_events']} inscripciones"
                        )

                # Botón para exportar a Excel
                if st.button("📊 Descargar Reporte Excel", type="primary"):
                    excel_data, filename = registration_system.export_team_report_to_excel(teams_data)
                    if excel_data:
                        st.download_button(
                            label="📥 Descargar Excel",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("¡Archivo Excel preparado para descarga!")
                    else:
                        st.error(f"Error generando Excel: {filename}")

        # Sección de Reporte de Medallas
        if swimmers:
            st.markdown("---")
            st.markdown("### 🏆 Reporte de Medallas")

            col1, col2 = st.columns(2)
            with col1:
                if st.button("👁️ Previsualizar Reporte de Medallas"):
                    with st.spinner("Generando previsualización..."):
                        medals_data, message = registration_system.preview_medals_report()
                        if medals_data:
                            st.session_state['medals_preview'] = medals_data
                            st.success("Previsualización generada")
                        else:
                            st.error(message)

            with col2:
                if st.button("🥇 Generar Reporte de Medallas"):
                    with st.spinner("Calculando medallas por categoría..."):
                        medals_data, message = registration_system.generate_medals_report()
                        if medals_data:
                            st.success(message)
                            st.session_state['medals_preview'] = medals_data
                            st.info("¡Reporte generado! Desplázate hacia abajo para descargar el Excel.")
                        else:
                            st.error(message)

            # Mostrar previsualización si está disponible
            if 'medals_preview' in st.session_state and st.session_state['medals_preview']:
                medals_data = st.session_state['medals_preview']

                # Mostrar resumen de medallas
                st.markdown("#### 🏆 Resumen de Medallas por Categoría")

                # Tabla de resumen
                summary_data = []
                for category, events in medals_data.items():
                    category_medals = {'Oro': 0, 'Plata': 0, 'Bronce': 0}
                    total_events = len(events)

                    for event_data in events.values():
                        medals = event_data['medals']
                        for medal_type, count in medals.items():
                            category_medals[medal_type] += count

                    summary_data.append({
                        'Categoría': category,
                        'Eventos': total_events,
                        'Oro': category_medals['Oro'],
                        'Plata': category_medals['Plata'],
                        'Bronce': category_medals['Bronce'],
                        'Total': sum(category_medals.values())
                    })

                import pandas as pd
                df_medals_summary = pd.DataFrame(summary_data)
                st.dataframe(df_medals_summary, use_container_width=True)

                # Botón para exportar medallas a Excel
                if st.button("🥇 Descargar Reporte Medallas Excel", type="primary"):
                    excel_data, filename = registration_system.export_medals_report_to_excel(medals_data)
                    if excel_data:
                        st.download_button(
                            label="📥 Descargar Excel Medallas",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("¡Archivo Excel de medallas preparado para descarga!")
                    else:
                        st.error(f"Error generando Excel: {filename}")

        # Sección de Reporte de Pagos
        if swimmers:
            st.markdown("---")
            st.markdown("### 💰 Reporte de Pagos de Clubes")

            col1, col2 = st.columns(2)
            with col1:
                if st.button("👁️ Previsualizar Reporte de Pagos"):
                    with st.spinner("Generando previsualización..."):
                        result = registration_system.preview_payments_report()
                        if len(result) == 4:
                            payments_data, swimmer_fee, team_fee, message = result
                            if payments_data:
                                st.session_state['payments_preview'] = (payments_data, swimmer_fee, team_fee)
                                st.success("Previsualización generada")
                            else:
                                st.error(message)
                        else:
                            st.error("Error en formato de respuesta")

            with col2:
                if st.button("💳 Generar Reporte de Pagos"):
                    with st.spinner("Calculando pagos por equipo..."):
                        result = registration_system.generate_payments_report()
                        if len(result) == 4:
                            payments_data, swimmer_fee, team_fee, message = result
                            if payments_data:
                                st.success(message)
                                st.session_state['payments_preview'] = (payments_data, swimmer_fee, team_fee)
                                st.info("¡Reporte generado! Desplázate hacia abajo para descargar el Excel.")
                            else:
                                st.error(message)
                        else:
                            st.error("Error en formato de respuesta")

            # Mostrar previsualización si está disponible
            if 'payments_preview' in st.session_state and st.session_state['payments_preview']:
                payments_data, swimmer_fee, team_fee = st.session_state['payments_preview']

                # Mostrar configuración de tarifas
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("💰 Valor por Nadador", f"${swimmer_fee:,.0f}")
                with col2:
                    st.metric("🏊‍♀️ Valor por Equipo", f"${team_fee:,.0f}")
                with col3:
                    total_revenue = sum(team_data['total_payment'] for team_data in payments_data.values())
                    st.metric("💵 Ingresos Totales", f"${total_revenue:,.0f}")

                # Tabla de resumen de pagos
                st.markdown("#### 💰 Resumen de Pagos por Equipo")
                summary_payments = []
                for team_name, team_data in payments_data.items():
                    summary_payments.append({
                        'Equipo': team_name,
                        'Nadadores': team_data['swimmer_count'],
                        'Pago Nadadores': f"${team_data['swimmer_fee_total']:,.0f}",
                        'Pago Equipo': f"${team_data['team_fee_total']:,.0f}",
                        'Total': f"${team_data['total_payment']:,.0f}"
                    })

                df_payments = pd.DataFrame(summary_payments)
                st.dataframe(df_payments, use_container_width=True)

                # Selector para generar cuenta de cobro por club
                st.markdown("#### 🧾 Generar Cuenta de Cobro por Club")
                team_names = list(payments_data.keys())
                if team_names:
                    selected_team_payment = st.selectbox(
                        "Selecciona el club para generar cuenta de cobro:",
                        team_names,
                        key="payment_team_selector"
                    )

                    col_invoice1, col_invoice2 = st.columns(2)
                    with col_invoice1:
                        if st.button(f"🧾 Generar Cuenta de Cobro - {selected_team_payment}", type="secondary"):
                            with st.spinner(f"Generando cuenta de cobro para {selected_team_payment}..."):
                                team_payment_data = payments_data[selected_team_payment]
                                pdf_data, filename = registration_system.generate_club_payment_invoice(
                                    selected_team_payment, team_payment_data, swimmer_fee, team_fee
                                )

                                if pdf_data:
                                    st.success(f"✅ Cuenta de cobro generada para {selected_team_payment}")
                                    st.download_button(
                                        label=f"⬇️ Descargar Cuenta de Cobro - {selected_team_payment}",
                                        data=pdf_data,
                                        file_name=filename,
                                        mime="application/pdf"
                                    )
                                else:
                                    st.error(f"❌ {filename}")

                    with col_invoice2:
                        # Mostrar resumen del equipo seleccionado para pagos
                        team_payment_info = payments_data[selected_team_payment]
                        st.info(
                            f"**{selected_team_payment}**\n\n"
                            f"👥 {team_payment_info['swimmer_count']} nadadores\n\n"
                            f"💰 Total: ${team_payment_info['total_payment']:,.0f}"
                        )

                # Botón para exportar pagos a Excel
                if st.button("💰 Descargar Reporte Pagos Excel", type="primary"):
                    excel_data, filename = registration_system.export_payments_report_to_excel(payments_data, swimmer_fee, team_fee)
                    if excel_data:
                        st.download_button(
                            label="📥 Descargar Excel Pagos",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("¡Archivo Excel de pagos preparado para descarga!")
                    else:
                        st.error(f"Error generando Excel: {filename}")
            
    
    # Copyright footer para todas las páginas
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666; font-size: 12px; margin-top: 20px;'>"
        "Sistema de Gestión de Competencias de Natación - Todos los derechos reservados"
        "</div>", 
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()