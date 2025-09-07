# generar_sembrado.py (Versión Corregida)

import pandas as pd
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# --- CONFIGURACIÓN ---
ARCHIVO_ENTRADA = 'planilla_inscripcion.xlsx' 
ARCHIVO_SALIDA = 'sembrado_competencia.xlsx'
CARRILES_PISCINA = 8

# --- LÓGICA DE PROCESAMIENTO ---

def parse_time(time_val):
    if pd.isna(time_val): return float('inf')
    if hasattr(time_val, 'minute'): return time_val.minute * 60 + time_val.second + time_val.microsecond / 1_000_000
    time_str = str(time_val).replace(',', '.')
    try:
        parts = time_str.split(':')
        if len(parts) == 2: return int(parts[0]) * 60 + float(parts[1])
        return float(time_str)
    except (ValueError, IndexError): return float('inf')

def seed_series(swimmers, lanes=8):
    lane_order = [4, 5, 3, 6, 2, 7, 1, 8]
    num_swimmers = len(swimmers)
    if num_swimmers == 0: return []
    num_series = math.ceil(num_swimmers / lanes)
    series_list = []
    sorted_swimmers = sorted(swimmers, key=lambda x: x['tiempo_en_segundos'])
    all_series_swimmers = []
    for i in range(num_series):
        start_index = max(0, num_swimmers - (i + 1) * lanes)
        end_index = num_swimmers - i * lanes
        serie_swimmers = sorted_swimmers[start_index:end_index]
        all_series_swimmers.append(serie_swimmers)
    for i, serie_swimmers in enumerate(all_series_swimmers, 1):
        serie_data = {"serie": i, "carriles": [None] * lanes}
        for j, swimmer in enumerate(serie_swimmers):
            target_lane = lane_order[j] - 1
            serie_data["carriles"][target_lane] = swimmer
        series_list.append(serie_data)
    return series_list

# --- FUNCIÓN PRINCIPAL DE EJECUCIÓN ---

def main():
    print("Iniciando sembrado por CATEGORÍA (versión corregida)...")
    try:
        df = pd.read_excel(ARCHIVO_ENTRADA)
        # Nombres de columna esperados
        info_cols = ['NOMBRE Y AP', 'EQUIPO', 'EDAD', 'CAT.', 'SEXO']
        event_cols = [col for col in df.columns if col not in info_cols and 'Nø' not in col and 'FECHA DE NA' not in col]
    except Exception as e:
        print(f"Error al leer el archivo de Excel '{ARCHIVO_ENTRADA}': {e}")
        return

    eventos = {}
    for index, row in df.iterrows():
        if pd.isna(row['NOMBRE Y AP']): continue
        sexo = row['SEXO'].upper()
        
        for prueba in event_cols:
            if pd.notna(row[prueba]):
                nombre_prueba = f"{prueba} - {'Mujeres' if sexo == 'F' else 'Hombres'}"
                if nombre_prueba not in eventos: eventos[nombre_prueba] = []
                
                nadador_info = {
                    "nombre": row['NOMBRE Y AP'], "equipo": row['EQUIPO'], "edad": int(row['EDAD']),
                    "categoria": row['CAT.'], # <-- ¡CORRECCIÓN CLAVE! Leemos la categoría directamente.
                    "tiempo_inscripcion": row[prueba], "tiempo_en_segundos": parse_time(row[prueba])
                }
                eventos[nombre_prueba].append(nadador_info)

    sembrado_final = {}
    for nombre_prueba, nadadores in eventos.items():
        nadadores_por_categoria = {}
        for nadador in nadadores:
            cat = nadador['categoria']
            if cat not in nadadores_por_categoria: nadadores_por_categoria[cat] = []
            nadadores_por_categoria[cat].append(nadador)
        
        sembrado_final[nombre_prueba] = []
        for categoria, lista_nadadores in sorted(nadadores_por_categoria.items()):
            series_sembradas = seed_series(lista_nadadores, CARRILES_PISCINA)
            sembrado_final[nombre_prueba].append({"categoria": categoria, "series": series_sembradas})

    wb = Workbook()
    ws = wb.active
    ws.title = "Sembrado por Pruebas"
    
    fila_actual = 1
    for nombre_prueba, categorias in sembrado_final.items():
        ws.cell(row=fila_actual, column=1, value=nombre_prueba).font = Font(bold=True, size=16)
        fila_actual += 2
        
        for data_categoria in categorias:
            ws.cell(row=fila_actual, column=1, value=f"Categoría: {data_categoria['categoria']}").font = Font(bold=True, size=14)
            fila_actual += 1
            
            for serie in data_categoria['series']:
                ws.cell(row=fila_actual, column=1, value=f"Serie {serie['serie']}").font = Font(bold=True)
                fila_actual += 1
                
                # Headers con columna adicional para tiempo de competencia
                headers = ["Carril", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo Inscripción", "Tiempo Competencia"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=fila_actual, column=col, value=header)
                    cell.font = Font(bold=True)
                    # Destacar la columna de tiempo de competencia
                    if header == "Tiempo Competencia":
                        cell.font = Font(bold=True, color="FF0000")  # Rojo para destacar
                fila_actual += 1
                
                for carril_num, nadador in enumerate(serie['carriles'], 1):
                    ws.cell(row=fila_actual, column=1, value=carril_num)
                    if nadador:
                        tiempo_val = nadador['tiempo_inscripcion']
                        tiempo_str = tiempo_val.strftime('%M:%S.%f')[:-4] if hasattr(tiempo_val, 'strftime') else str(tiempo_val)
                        ws.cell(row=fila_actual, column=2, value=nadador['nombre'])
                        ws.cell(row=fila_actual, column=3, value=nadador['equipo'])
                        ws.cell(row=fila_actual, column=4, value=nadador['edad'])
                        ws.cell(row=fila_actual, column=5, value=nadador['categoria'])           
                        ws.cell(row=fila_actual, column=6, value=tiempo_str)
                        # Columna vacía para tiempo de competencia (editable)
                        comp_cell = ws.cell(row=fila_actual, column=7, value="")
                        comp_cell.font = Font(color="0000FF")  # Azul para indicar que es editable
                    fila_actual += 1
                fila_actual += 1
    
    ws.column_dimensions['B'].width = 40
    wb.save(ARCHIVO_SALIDA)
    print(f"¡Éxito! Archivo '{ARCHIVO_SALIDA}' generado con la columna 'Categoría'.")

def get_seeding_data():
    """Retorna los datos del sembrado para visualización sin generar archivo"""
    try:
        df = pd.read_excel(ARCHIVO_ENTRADA)
        info_cols = ['NOMBRE Y AP', 'EQUIPO', 'EDAD', 'CAT.', 'SEXO']
        event_cols = [col for col in df.columns if col not in info_cols and 'Nø' not in col and 'FECHA DE NA' not in col]
    except Exception as e:
        return None, f"Error al leer el archivo de Excel: {e}"

    eventos = {}
    for index, row in df.iterrows():
        if pd.isna(row['NOMBRE Y AP']): continue
        sexo = row['SEXO'].upper()
        categoria = row['CAT.']
        
        for prueba in event_cols:
            if pd.notna(row[prueba]):
                nombre_prueba = f"{prueba} - {categoria} - {'Mujeres' if sexo == 'F' else 'Hombres'}"
                if nombre_prueba not in eventos: eventos[nombre_prueba] = []
                
                nadador_info = {
                    "nombre": row['NOMBRE Y AP'], "equipo": row['EQUIPO'], "edad": int(row['EDAD']),
                    "categoria": row['CAT.'],
                    "tiempo_inscripcion": row[prueba], "tiempo_en_segundos": parse_time(row[prueba])
                }
                eventos[nombre_prueba].append(nadador_info)

    sembrado_final = {}
    for nombre_prueba, nadadores in eventos.items():
        sembrado_final[nombre_prueba] = {"series": seed_series(nadadores, CARRILES_PISCINA)}

    return sembrado_final, "Sembrado generado exitosamente"

def main_full():
    """Función completa para usar desde app.py"""
    main()

if __name__ == "__main__":
    main()