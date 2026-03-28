#!/usr/bin/env python3
"""
Procesador de Sembrado con Tiempos de Competencia
Convierte archivos de sembrado con tiempos agregados en formato de resultados
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def _parse_seeding_sheet_df(df_raw, resultados_data):
    """Recorre una hoja de sembrado (formato estándar) y agrega filas con tiempo de competencia."""
    current_event = None
    current_series = None

    for index, row in df_raw.iterrows():
        row_values = [str(val) if pd.notna(val) else "" for val in row.values]

        if len(row_values) > 0 and " - " in row_values[0] and row_values[0] != "":
            if all(val == "" or val == "nan" for val in row_values[1:7]):
                current_event = row_values[0]
                continue

        if len(row_values) > 0 and row_values[0].startswith("Serie "):
            try:
                current_series = int(row_values[0].split()[1])
                continue
            except (ValueError, IndexError):
                pass

        if len(row_values) > 0 and row_values[0] == "Carril":
            continue

        try:
            carril = int(float(row_values[0]))
            nombre = row_values[1] if len(row_values) > 1 else ""
            equipo = row_values[2] if len(row_values) > 2 else ""
            edad = row_values[3] if len(row_values) > 3 else ""
            categoria = row_values[4] if len(row_values) > 4 else ""
            tiempo_inscripcion = row_values[5] if len(row_values) > 5 else ""
            tiempo_competencia = row_values[6] if len(row_values) > 6 else ""

            if nombre and nombre != "---" and tiempo_competencia and tiempo_competencia.strip() != "":
                resultados_data.append({
                    'Evento': current_event or "Evento sin nombre",
                    'Nombre': nombre,
                    'Equipo': equipo,
                    'Edad': edad,
                    'Categoria': categoria,
                    'Tiempo': tiempo_competencia,
                    'Serie': current_series or 1,
                    'Carril': carril
                })

        except (ValueError, TypeError):
            continue

def process_seeding_with_times(input_file):
    """
    Procesa un archivo de sembrado que contiene tiempos de competencia
    y genera un archivo de resultados estructurado
    """
    if not os.path.exists(input_file):
        return False, f"Archivo no encontrado: {input_file}"
    
    try:
        resultados_data = []
        xl = pd.ExcelFile(input_file)
        for sheet_name in xl.sheet_names:
            df_raw = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
            _parse_seeding_sheet_df(df_raw, resultados_data)

        if len(resultados_data) == 0:
            return False, "No se encontraron tiempos de competencia válidos en el archivo"
        
        # Crear archivo de resultados
        output_file = f"resultados_desde_sembrado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados de Competencia"
        
        # Headers
        headers = ["Evento", "Nombre", "Equipo", "Edad", "Categoría", "Tiempo", "Serie", "Carril"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Escribir datos
        for row_idx, data in enumerate(resultados_data, 2):
            ws.cell(row=row_idx, column=1, value=data['Evento'])
            ws.cell(row=row_idx, column=2, value=data['Nombre'])
            ws.cell(row=row_idx, column=3, value=data['Equipo'])
            ws.cell(row=row_idx, column=4, value=data['Edad'])
            ws.cell(row=row_idx, column=5, value=data['Categoria'])
            ws.cell(row=row_idx, column=6, value=data['Tiempo'])
            ws.cell(row=row_idx, column=7, value=data['Serie'])
            ws.cell(row=row_idx, column=8, value=data['Carril'])
        
        # Ajustar ancho de columnas
        column_widths = {'A': 40, 'B': 30, 'C': 20, 'D': 8, 'E': 12, 'F': 15, 'G': 8, 'H': 8}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(output_file)
        
        return True, f"Archivo de resultados generado: {output_file} con {len(resultados_data)} registros"
        
    except Exception as e:
        return False, f"Error al procesar archivo: {str(e)}"

def main():
    """Función principal para uso desde línea de comandos"""
    print("=== PROCESADOR DE SEMBRADO CON TIEMPOS ===")
    
    # Buscar archivos de sembrado
    seeding_files = []
    if os.path.exists("sembrado_competencia.xlsx"):
        seeding_files.append("sembrado_competencia.xlsx")
    if os.path.exists("sembrado_competencia_POR_TIEMPO.xlsx"):
        seeding_files.append("sembrado_competencia_POR_TIEMPO.xlsx")
    
    if not seeding_files:
        print("❌ No se encontraron archivos de sembrado")
        print("💡 Primero genera un sembrado y agrega los tiempos de competencia")
        return
    
    print("\n📁 Archivos de sembrado encontrados:")
    for i, file in enumerate(seeding_files, 1):
        print(f"   {i}. {file}")
    
    try:
        choice = input(f"\n🔸 Selecciona archivo (1-{len(seeding_files)}): ")
        selected_file = seeding_files[int(choice) - 1]
        
        print(f"\n🔄 Procesando {selected_file}...")
        success, message = process_seeding_with_times(selected_file)
        
        if success:
            print(f"✅ {message}")
        else:
            print(f"❌ {message}")
            
    except (ValueError, IndexError):
        print("❌ Selección inválida")
    except KeyboardInterrupt:
        print("\n👋 Operación cancelada")

if __name__ == "__main__":
    main()