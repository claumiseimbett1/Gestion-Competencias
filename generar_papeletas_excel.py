# generar_papeletas_excel.py
import pandas as pd
import os
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN ---
ARCHIVO_PAPELETAS_EXCEL = 'papeletas_jueces.xlsx'

def parse_time(time_val):
    """Convierte tiempo a segundos para ordenamiento"""
    if pd.isna(time_val): return float('inf')
    if hasattr(time_val, 'minute'): return time_val.minute * 60 + time_val.second + time_val.microsecond / 1_000_000
    time_str = str(time_val).replace(',', '.')
    try:
        parts = time_str.split(':')
        if len(parts) == 2: return int(parts[0]) * 60 + float(parts[1])
        return float(time_str)
    except (ValueError, IndexError): return float('inf')

def leer_datos_sembrado():
    """Lee los datos del sembrado con series y carriles asignados"""
    try:
        df = pd.read_excel('planilla_inscripcion.xlsx')
        info_cols = ['NOMBRE Y AP', 'EQUIPO', 'EDAD', 'CAT.', 'SEXO']
        event_cols = [col for col in df.columns if col not in info_cols and 'Nø' not in col and 'FECHA DE NA' not in col]
        
        eventos = {}
        for index, row in df.iterrows():
            if pd.isna(row['NOMBRE Y AP']): continue
            sexo = row['SEXO'].upper()
            
            for prueba in event_cols:
                if pd.notna(row[prueba]):
                    nombre_prueba = f"{prueba} - {'Mujeres' if sexo == 'F' else 'Hombres'}"
                    if nombre_prueba not in eventos: eventos[nombre_prueba] = []
                    
                    nadador_info = {
                        "nombre": row['NOMBRE Y AP'], 
                        "equipo": row['EQUIPO'], 
                        "edad": int(row['EDAD']),
                        "categoria": row['CAT.'],
                        "sexo": sexo,
                        "tiempo_inscripcion": row[prueba], 
                        "tiempo_en_segundos": parse_time(row[prueba])
                    }
                    eventos[nombre_prueba].append(nadador_info)
        
        # Agrupar por categoría y crear series con carriles asignados
        papeletas_con_carriles = []
        for nombre_prueba, nadadores in eventos.items():
            nadadores_por_categoria = {}
            for nadador in nadadores:
                cat = nadador['categoria']
                if cat not in nadadores_por_categoria: 
                    nadadores_por_categoria[cat] = []
                nadadores_por_categoria[cat].append(nadador)
            
            for categoria, lista_nadadores in sorted(nadadores_por_categoria.items()):
                # Ordenar por tiempo
                lista_nadadores_ordenada = sorted(lista_nadadores, key=lambda x: x['tiempo_en_segundos'])
                
                # Crear series de 8 nadadores
                num_nadadores = len(lista_nadadores_ordenada)
                num_series = math.ceil(num_nadadores / 8)
                
                for serie_num in range(1, num_series + 1):
                    inicio = (serie_num - 1) * 8
                    fin = min(serie_num * 8, num_nadadores)
                    nadadores_serie = lista_nadadores_ordenada[inicio:fin]
                    
                    # Asignar carriles usando el orden estándar
                    lane_order = [4, 5, 3, 6, 2, 7, 1, 8]
                    
                    for i, nadador in enumerate(nadadores_serie):
                        if i < len(lane_order):
                            carril_asignado = lane_order[i]
                            
                            papeleta = {
                                "nombre": nadador['nombre'],
                                "equipo": nadador['equipo'],
                                "categoria": nadador['categoria'],
                                "sexo": nadador['sexo'],
                                "prueba": nombre_prueba,
                                "serie": serie_num,
                                "carril": carril_asignado,
                                "tiempo_inscripcion": nadador['tiempo_inscripcion']
                            }
                            papeletas_con_carriles.append(papeleta)
        
        return papeletas_con_carriles
    
    except Exception as e:
        print(f"Error al leer datos del sembrado: {e}")
        return []

def generar_papeletas_excel():
    """Genera papeletas con datos completos del sembrado, 3 por hoja"""
    papeletas_sembrado = leer_datos_sembrado()
    
    if not papeletas_sembrado:
        return False, "No se pudieron leer los datos del sembrado"
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Papeletas Jueces"
        
        # Configurar anchos de columna para dar espacio
        for col in range(1, 9):  # A hasta H
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 3 papeletas por hoja
        fila_actual = 1
        
        # Estilos
        titulo_font = Font(bold=True, size=12, color='1E88E5')
        info_font = Font(size=9)
        serie_font = Font(bold=True, size=10)
        tiempo_font = Font(bold=True, size=14, color='FF0000')
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        border_thick = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'), 
            bottom=Side(style='thick')
        )
        
        # Configurar anchos de columna para mejor distribución (3 papeletas por fila)
        for col in range(1, 10):  # A hasta I (3 papeletas × 3 columnas cada una)
            ws.column_dimensions[get_column_letter(col)].width = 11
        
        # Agregar título del documento
        ws.merge_cells('A1:I1')
        title_cell = ws.cell(row=1, column=1, value="PAPELETAS DE JUECES - COMPETENCIA DE NATACIÓN TEN")
        title_cell.font = Font(bold=True, size=14, color='1E88E5')
        title_cell.alignment = Alignment(horizontal='center')
        
        fila_actual = 3  # Empezar después del título
        contador_papeletas = 0
        
        for papeleta in papeletas_sembrado:
            # Calcular posición (3 papeletas por fila)
            pos_en_fila = contador_papeletas % 3  # 0, 1, 2
            if pos_en_fila == 0 and contador_papeletas > 0:
                fila_actual += 8  # Espacio entre filas de papeletas (reducido)
            
            col_inicio = pos_en_fila * 3 + 1  # 1, 4, 7
            
            fila_base = fila_actual
            
            # PRIMERA FILA: Prueba
            ws.merge_cells(start_row=fila_base, start_column=col_inicio, 
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio, value=f"PRUEBA: {papeleta['prueba']}")
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border_thick
            
            # SEGUNDA FILA: Nadador, Equipo, Categoría  
            fila_base += 1
            ws.merge_cells(start_row=fila_base, start_column=col_inicio,
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio,
                         value=f"{papeleta['nombre']}\n{papeleta['equipo']} - {papeleta['categoria']}")
            cell.font = info_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border_thin
            ws.row_dimensions[fila_base].height = 30
            
            # TERCERA FILA: Serie y Carril YA ASIGNADOS
            fila_base += 1
            ws.merge_cells(start_row=fila_base, start_column=col_inicio,
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio, 
                         value=f"SERIE: {papeleta['serie']}  |  CARRIL: {papeleta['carril']}")
            cell.font = serie_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            
            # CUARTA FILA: TIEMPO DE COMPETENCIA (título)
            fila_base += 1
            ws.merge_cells(start_row=fila_base, start_column=col_inicio,
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio, value="TIEMPO DE COMPETENCIA:")
            cell.font = tiempo_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            
            # QUINTA FILA: Línea para anotación (EN BLANCO) - más compacta
            fila_base += 1
            ws.merge_cells(start_row=fila_base, start_column=col_inicio,
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio, value="_____ : _____ . _____")
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thick
            
            # SEXTA FILA: Juez (más compacta)
            fila_base += 1
            ws.merge_cells(start_row=fila_base, start_column=col_inicio,
                         end_row=fila_base, end_column=col_inicio + 2)
            cell = ws.cell(row=fila_base, column=col_inicio, value="Juez: ___________________")
            cell.font = Font(size=8, color='666666')
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            
            # Ajustar altura de las filas
            ws.row_dimensions[fila_base - 1].height = 25  # Fila de tiempo
            ws.row_dimensions[fila_base].height = 15      # Fila de juez
            
            contador_papeletas += 1
        
        # Configurar márgenes de página para impresión optimizada
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4 
        ws.page_margins.bottom = 0.3
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
        
        # Configurar orientación horizontal para mejor aprovechamiento
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # Configurar escala de impresión para que quepa todo
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Sin límite de altura
        
        # Configurar repetir filas en la parte superior (título)
        ws.print_title_rows = '1:2'
        
        # Configurar líneas de cuadrícula para impresión
        ws.print_options.gridLines = True
        ws.print_options.gridLinesSet = True
        
        wb.save(ARCHIVO_PAPELETAS_EXCEL)
        return True, f"Papeletas Excel generadas exitosamente: {ARCHIVO_PAPELETAS_EXCEL}"
        
    except Exception as e:
        return False, f"Error al generar papeletas Excel: {e}"

def main():
    """Función principal para ejecutar desde línea de comandos"""
    print("Generando papeletas Excel para jueces...")
    
    if not os.path.exists('planilla_inscripcion.xlsx'):
        print("ERROR: No se encontro el archivo 'planilla_inscripcion.xlsx'")
        print("Este archivo es necesario para generar las papeletas.")
        return
    
    success, message = generar_papeletas_excel()
    print(message)
    
    if success:
        total_papeletas = len(leer_datos_sembrado())
        print(f"Total de papeletas generadas: {total_papeletas}")
        print(f"Archivo guardado como: {ARCHIVO_PAPELETAS_EXCEL}")
        print(f"Formato: 3 papeletas por hoja, con serie y carril asignados")

if __name__ == "__main__":
    main()